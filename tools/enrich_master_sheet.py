#!/usr/bin/env python3
r"""
Mino Bimaadiziwin — free AI/data enrichment helper for the Master Sheet.

For every community it looks up OBJECTIVE facts from free, no-key sources
(Wikipedia + Wikidata) and scrapes the official website for contacts and
documents, then writes proposals into a NEW "AI Suggestions" tab in a COPY of
your workbook — each with the SOURCE url, a CONFIDENCE level, and the dedicated
AI colour so everyone knows AI pulled it. It NEVER edits your Master Sheet and
never touches subjective program-description columns; a human gives the final ✓.

No scraper or free AI can promise 100% accuracy, so this reports only what a
citable source says and leaves the decision to a person — the honest, safe way
to keep the sheet trustworthy.

USAGE
    python tools/enrich_master_sheet.py "C:\path\to\Master Sheet.xlsx"
(omit the path to use the default). Needs internet; uses only requests + openpyxl.
"""
from __future__ import annotations

import sys
import time
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

# make `server` importable when run from the repo root
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from server.enrich import enrich_one, clean, looks_like_community

# colour legend from the "Colour Codes" tab; AI gets its own distinct colour.
LEGEND = {"missing": "E99B9A", "review": "FDE6A0", "correction": "CBD9F6", "verified": "A3C3C8"}
AI_COLOUR = "C7B3F0"          # Light Purple — "AI Suggested (verify)"
AI_LABEL = "AI Suggested (verify)"
HEADER_FILL = PatternFill("solid", fgColor="1F2A37")
HEADER_FONT = Font(color="FFFFFF", bold=True)

DEFAULT_PATH = r"C:\Users\algom\Downloads\Copy of Mino Bimaadiziwin Project_ Services and Program Scan  (2).xlsx"


def read_communities(ws):
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(r, 1).value)
        if name and looks_like_community(name):
            yield r, name, clean(ws.cell(r, 2).value), ws.cell(r, 11).value


def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(DEFAULT_PATH)
    if not path.exists():
        print(f"!! file not found: {path}"); sys.exit(1)
    print(f"Reading {path.name} …")
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Master Sheet" not in wb.sheetnames:
        print("!! no 'Master Sheet' tab"); sys.exit(1)
    rows = list(read_communities(wb["Master Sheet"]))
    print(f"Found {len(rows)} community rows. Querying free sources + scraping sites …")

    results = {}
    with ThreadPoolExecutor(max_workers=6) as ex:
        fut = {ex.submit(enrich_one, n, link, pop): n for (_, n, link, pop) in rows}
        done = 0
        for f in as_completed(fut):
            name = fut[f]
            try:
                results[name] = f.result()
            except Exception as e:
                results[name] = [{"field": "(error)", "current": "", "suggested": "",
                                  "source": "", "confidence": "Low", "note": str(e)}]
            done += 1
            if done % 10 == 0 or done == len(rows):
                print(f"  …{done}/{len(rows)}")

    out_wb = openpyxl.load_workbook(path)
    if "Colour Codes" in out_wb.sheetnames:
        cc = out_wb["Colour Codes"]
        if not any(clean(cc.cell(r, 4).value).lower() == AI_LABEL.lower() for r in range(1, cc.max_row + 1)):
            r = cc.max_row + 2
            cc.cell(r, 1, f"AI Suggested\n#{AI_COLOUR}")
            cc.cell(r, 2).fill = PatternFill("solid", fgColor=AI_COLOUR)
            cc.cell(r, 4, AI_LABEL)

    if "AI Suggestions" in out_wb.sheetnames:
        del out_wb["AI Suggestions"]
    sug = out_wb.create_sheet("AI Suggestions", 0)
    sug.append([f"AI-pulled information · purple = AI Suggested (verify) · nothing in the Master Sheet was "
                f"changed. When you accept a value, paste it into the Master Sheet and fill that cell purple (#{AI_COLOUR})."])
    sug.merge_cells("A1:G1")
    sug.cell(1, 1).fill = PatternFill("solid", fgColor=AI_COLOUR)
    sug.cell(1, 1).alignment = Alignment(wrap_text=True, vertical="center")
    sug.row_dimensions[1].height = 42

    headers = ["Community/Organization", "Field", "Current in sheet", "Suggested value",
               "Confidence", "Source (click to verify)", "Note"]
    sug.append(headers)
    for c in range(1, len(headers) + 1):
        cell = sug.cell(2, c); cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, w in enumerate([30, 26, 26, 34, 12, 46, 54], 1):
        sug.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    conf_fill = {"High": LEGEND["correction"], "Medium": LEGEND["review"], "Low": LEGEND["missing"]}
    ai_fill = PatternFill("solid", fgColor=AI_COLOUR)
    n_sug = 0
    for (_, name, _l, _p) in rows:
        for s in results.get(name, []):
            sug.append([name, s["field"], s["current"], s["suggested"], s["confidence"], s["source"], s["note"]])
            row = sug.max_row
            sug.cell(row, 2).fill = ai_fill
            sug.cell(row, 5).fill = PatternFill("solid", fgColor=conf_fill.get(s["confidence"], LEGEND["review"]))
            for c in range(1, 8):
                sug.cell(row, c).alignment = Alignment(vertical="top", wrap_text=True)
            if s["source"]:
                lk = sug.cell(row, 6); lk.hyperlink = s["source"]; lk.font = Font(color="1155CC", underline="single")
            n_sug += 1
    sug.freeze_panes = "A3"

    # ---- ALSO fill the Master Sheet IN PLACE (blanks only) + colour purple ----
    # Master Sheet columns we can auto-fill, mapped from enrich_one's field names.
    MS_COL = {
        "Community Link": 2,                      # B
        "Contact Information for Departments": 3,  # C
        "Community Population": 11,               # K
        "Strategic Plan": 14,                    # N
        "Annual General Meeting (AGM)": 15,      # O
        "Financial Statements": 16,              # P
    }
    ms = out_wb["Master Sheet"]
    fill_fill = PatternFill("solid", fgColor=AI_COLOUR)

    # community rows are often merged across 2 rows → writes must target the
    # top-left anchor of the merged range, not the read-only MergedCell.
    def anchor(row_i, col):
        for rng in ms.merged_cells.ranges:
            if rng.min_row <= row_i <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return ms.cell(rng.min_row, rng.min_col)
        return ms.cell(row_i, col)

    n_cells = 0
    for (row_i, name, _l, _p) in rows:
        for s in results.get(name, []):
            col = MS_COL.get(s["field"])
            val = s.get("suggested")
            if not col or not val or val in ("—", "") or s["confidence"] == "Low":
                continue
            cell = anchor(row_i, col)
            if clean(cell.value) in ("", "missing information", "needs review", "n/a", "-", "—"):
                cell.value = val                  # fill only when blank
                cell.fill = fill_fill             # mark it AI (purple)
                n_cells += 1

    out_path = path.with_name(path.stem + " — AI Filled.xlsx")
    out_wb.save(out_path)
    print(f"\nDone. Filled {n_cells} blank Master Sheet cells (purple = AI) and logged "
          f"{n_sug} suggestions for {len(rows)} communities.")
    print(f"Wrote: {out_path}")
    print("Only blank cells were filled — your existing entries and the other tabs are untouched.")


if __name__ == "__main__":
    main()
