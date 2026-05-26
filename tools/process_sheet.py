#!/usr/bin/env python3
"""
Mino Bimaadiziwin Atlas — Sheet Processor (v2)
==============================================
Reads an Excel workbook OR a CSV file and produces `communities-data.js` for
the front-end. Run this whenever the team updates the master sheet.

USAGE
-----
    # Install deps once  (openpyxl needed only for .xlsx)
    pip install openpyxl

    # Excel
    python tools/process_sheet.py path/to/master.xlsx

    # CSV (UTF-8, comma- or tab-delimited — auto-detected)
    python tools/process_sheet.py path/to/master.csv

    # Custom output and/or sheet
    python tools/process_sheet.py master.xlsx --out ../communities-data.js --sheet "Master Sheet"

WHAT'S NEW IN v2
----------------
* Accepts BOTH .xlsx and .csv (auto-detects from extension; tab/comma sniff).
* HEADER-DRIVEN column mapping: looks at the first 10 rows for a header
  containing words like "name", "physical health", "contact info", "AGM",
  "population", etc. and maps them to canonical fields. No more brittle
  fixed-index lookup — if a column is missing or reordered, the script
  still works.
* Each missing column is reported once at the top of the run so the team
  can see exactly what didn't land.
* Section / direction / org-type classification is unchanged (still
  inferred from the row context) — but now also from explicit
  "Direction" / "Type" / "Region" columns if your sheet has them.
* Preserves lat/lon from the previous communities-data.js by normalized
  name, exactly as before.
* Reports parsing diagnostics: rows skipped, rows merged as duplicates,
  records with no contact info, records with no narrative content, etc.

The output is what `helpers.jsx` consumes, so the dashboard live-updates
on next page load.
"""

from __future__ import annotations
import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Any, Iterable

# openpyxl is only needed for .xlsx files; import lazily.
def _load_openpyxl():
    try:
        from openpyxl import load_workbook
        return load_workbook
    except ImportError:
        sys.exit("Missing dependency for .xlsx: pip install openpyxl")


# ---------------------------------------------------------------------------
# Canonical fields and the header-name patterns that map to them.
# Each canonical field has a list of regex patterns; the FIRST column whose
# header matches one of these patterns wins. Case- and whitespace-insensitive.
# ---------------------------------------------------------------------------
CANONICAL_FIELDS: dict[str, list[str]] = {
    'name':            [r'^name$', r'community.*name', r'organi[sz]ation.*name', r'^community$', r'^first nation$'],
    'link':            [r'^link', r'^website', r'^url', r'^homepage'],
    'contact':         [r'contact', r'staff', r'phone.*list', r'directory'],
    'physical':        [r'physical.*health', r'^physical', r'medical', r'clinic'],
    'mental':          [r'mental.*health', r'^mental', r'counsell?ing', r'therapy'],
    'spiritual':       [r'spiritual', r'ceremon', r'cultural.*support'],
    'emotional':       [r'emotional', r'wellness'],
    'survivors':       [r'survivor', r'residential.*school', r'sixties.*scoop', r'irs'],
    'youth':           [r'youth', r'young.*people', r'children'],
    'connect':         [r'connect', r'intergenerational', r'elders?.*youth'],
    'population':      [r'^population$', r'members?$', r'on.?reserve', r'band.*member'],
    'populationNote':  [r'population.*note', r'population.*detail', r'pop.*source'],
    'additionalLinks': [r'additional.*link', r'other.*link', r'social.*media'],
    'strategicPlan':   [r'strategic.*plan', r'^plan$', r'health.*plan'],
    'agm':             [r'^agm', r'annual.*general', r'annual.*report'],
    'financials':      [r'financ', r'audit', r'budget'],
    'notes':           [r'^notes?$', r'comment', r'observation'],
    # Optional explicit overrides — if your sheet has these columns, we trust them
    'direction':       [r'^direction$', r'sacred.*direction', r'^region.*group'],
    'type':            [r'^type$', r'organi[sz]ation.*type', r'^category$'],
    'lat':             [r'^lat(itude)?$'],
    'lon':             [r'^lon(g(itude)?)?$'],
}

# Default index-based fallback layout (matches the original master sheet).
# Used ONLY if no header row is detected. Keys are canonical field names.
LEGACY_INDEX_LAYOUT: dict[str, int] = {
    'name': 0, 'link': 1, 'contact': 2,
    'physical': 3, 'mental': 4, 'spiritual': 5, 'emotional': 6,
    'survivors': 7, 'youth': 8, 'connect': 9,
    'population': 10, 'populationNote': 11, 'additionalLinks': 12,
    'strategicPlan': 13, 'agm': 14, 'financials': 15,
    'notes': 17,
}

# Section headers (rows where only column A is filled with a known label).
SECTION_NAMES = {
    'first 10 (pilot)',
    'second project - environmental scan - local and regional communities',
    'algoma district first nation communities',
    'east', 'south', 'west', 'north',
    'indigenous health authorities',
    'indigenous organizations',
    'indigenous friendship centres',
    'child and family services',
    'umbrella organizations',
    'nishnawbe aski nation', 'mushkegowuk council', 'james bay cree',
    'access open minds\n(esprits ouverts)',
    'access open minds (esprits ouverts)',
}


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------
def _slug(s: str) -> str:
    return re.sub(r'[^a-z0-9 ]', ' ', (s or '').lower()).strip()


def detect_header(rows: list[list[str]], max_scan: int = 10) -> tuple[int, dict[str, int]]:
    """Return (header_row_index, column_map). column_map maps canonical
    field name -> source column index. If no header is detected, returns
    (-1, LEGACY_INDEX_LAYOUT)."""
    best = (-1, 0, {})  # row, hit_count, mapping
    for i, row in enumerate(rows[:max_scan]):
        mapping: dict[str, int] = {}
        seen_cells = 0
        for col_idx, cell in enumerate(row):
            text = _slug(str(cell))
            if not text or len(text) > 80:
                continue
            seen_cells += 1
            for field, patterns in CANONICAL_FIELDS.items():
                if field in mapping:
                    continue
                for pat in patterns:
                    if re.search(pat, text):
                        mapping[field] = col_idx
                        break
        hits = len(mapping)
        # A "real" header should have at least name + 4 other canonical fields
        if hits >= 5 and 'name' in mapping and hits > best[1]:
            best = (i, hits, mapping)
    if best[0] == -1:
        return -1, dict(LEGACY_INDEX_LAYOUT)
    return best[0], best[2]


def cell(row: list[str], col_map: dict[str, int], key: str) -> str:
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return ''
    val = row[idx]
    return '' if val is None else str(val).strip()


# ---------------------------------------------------------------------------
# Section / classification rules
# ---------------------------------------------------------------------------
def normalize(s: str) -> str:
    s = (s or '').lower()
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r"[\s\(\)\/,'\-\.#]", '', s)
    s = re.sub(r'firstnations?', 'fn', s)
    return s.strip()


def is_section_row(cells: list[str]) -> bool:
    """A section header is a row with ONLY column A filled, and the value is a
    known section name. Be strict: a short A-only value (like 'test1') is a
    WIP community record, not a section header. We'd rather show a sparse
    row in the dashboard than silently drop the team's data.
    """
    a = (cells[0] or '').strip().lower()
    other = sum(1 for c in cells[1:] if (c or '').strip())
    if other > 0:
        return False
    return a in SECTION_NAMES


def classify(rec: dict) -> dict:
    # Trust explicit columns if present
    if rec.get('type'):
        rec['type'] = rec['type'].strip()
    else:
        sec = (rec.get('section') or '').lower()
        if 'health authorit' in sec:
            rec['type'] = 'Health Authority'
        elif 'friendship' in sec:
            rec['type'] = 'Friendship Centre'
        elif 'child and family' in sec or 'child & family' in sec:
            rec['type'] = 'Child & Family Services'
        elif 'umbrella' in sec:
            rec['type'] = 'Umbrella Organization'
        elif 'organization' in sec and 'umbrella' not in sec:
            rec['type'] = 'Indigenous Organization'
        else:
            rec['type'] = 'Community'

    if not rec.get('direction'):
        sec = (rec.get('section') or '').lower()
        if rec.get('section') in ('East', 'South', 'West', 'North'):
            rec['direction'] = rec['section']
        elif 'algoma' in sec or 'pilot' in sec:
            rec['direction'] = 'North'
        elif rec['type'] == 'Community':
            rec['direction'] = 'North'
        else:
            rec['direction'] = 'Central'
    return rec


# ---------------------------------------------------------------------------
# Contact parsing (unchanged)
# ---------------------------------------------------------------------------
NAME_RX = re.compile(r"^([A-Z][a-z]+(?:[\s\-'][A-Z][a-z]+){1,3})")
ROLE_KEYWORDS = re.compile(
    r'coordinator|director|manager|admin|chief|councillor|councilor|nurse|'
    r'worker|assistant|officer|specialist|supervisor|lead|head|advisor|liaison|'
    r'representative|practitioner|therapist|youth|elder|wellness|health|reception',
    re.I,
)
PHONE_RX = re.compile(
    r'(?:\(?\d{3}\)?[\s\-\.]?\d{3}[\s\-\.]?\d{4})|'
    r'1[-\s]?\d{3}[-\s]?\d{3}[-\s]?\d{4}'
)
EMAIL_RX = re.compile(r'[\w.+-]+@[\w-]+\.[\w.-]+')
EXT_RX = re.compile(r'\bext\.?\s*:?\s*(\d+)', re.I)


def parse_contacts(text: str) -> list[dict]:
    if not text:
        return []
    t = text.replace('\r', '')
    blocks = re.split(r'\n{2,}', t)
    staff = []
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        lines = [l.strip() for l in block.split('\n') if l.strip()]
        name = role = phone = email = ext = ''
        for line in lines:
            m = NAME_RX.match(line)
            if m and not name and '@' not in line:
                name = m.group(1)
            if not role and ROLE_KEYWORDS.search(line) \
               and '@' not in line and not PHONE_RX.search(line) \
               and line != name:
                role = line[:120]
            p = PHONE_RX.search(line)
            if p and not phone:
                phone = p.group(0)
            e = EMAIL_RX.search(line)
            if e and not email:
                email = e.group(0)
            x = EXT_RX.search(line)
            if x:
                ext = x.group(1)
        if name or phone or email:
            staff.append({'name': name, 'role': role,
                          'phone': phone, 'ext': ext, 'email': email,
                          'raw': block[:400]})
    return staff


# ---------------------------------------------------------------------------
# Row sources — Excel + CSV  → list[list[str]]
# ---------------------------------------------------------------------------
def load_xlsx_rows(path: Path, sheet_name: str | None) -> tuple[list[list[str]], str]:
    load_workbook = _load_openpyxl()
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    names = wb.sheetnames
    target = sheet_name or ('Master Sheet' if 'Master Sheet' in names else names[0])
    if target not in names:
        sys.exit(f'Sheet "{target}" not found. Available: {", ".join(names)}')
    ws = wb[target]
    rows = []
    for raw in ws.iter_rows(values_only=True):
        rows.append(['' if v is None else str(v) for v in (raw or ())])
    return rows, target


def load_csv_rows(path: Path) -> tuple[list[list[str]], str]:
    text = path.read_text(encoding='utf-8-sig', errors='replace')
    # Sniff delimiter
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
    except csv.Error:
        class D: delimiter = ','; quotechar = '"'; doublequote = True; skipinitialspace = False
        dialect = D()
    reader = csv.reader(text.splitlines(), dialect=dialect)
    rows = [['' if v is None else str(v) for v in r] for r in reader]
    return rows, path.name


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------
def process_rows(rows: list[list[str]], previous: list[dict] | None = None,
                 verbose: bool = True) -> list[dict]:
    if not rows:
        return []

    # 1. Detect header
    hdr_idx, col_map = detect_header(rows)
    if verbose:
        if hdr_idx >= 0:
            print(f'[header] found at row {hdr_idx + 1}; '
                  f'mapped {len(col_map)} columns: {", ".join(col_map.keys())}',
                  file=sys.stderr)
            missing = [k for k in CANONICAL_FIELDS if k not in col_map
                       and k not in ('direction', 'type', 'lat', 'lon',
                                      'populationNote', 'notes')]
            if missing:
                print(f'[header] missing optional columns (will be blank): '
                      f'{", ".join(missing)}', file=sys.stderr)
        else:
            print('[header] none detected → falling back to legacy fixed-index layout',
                  file=sys.stderr)

    # 2. Walk rows, skipping section headers and junk
    cur_section = ''
    records: list[dict] = []
    dedup: dict[str, int] = {}
    stats = {'skipped_junk': 0, 'skipped_dup_marker': 0, 'merged_dupes': 0,
             'section_rows': 0, 'no_contact': 0, 'no_narrative': 0}

    start = hdr_idx + 1 if hdr_idx >= 0 else 0
    for i in range(start, len(rows)):
        raw = rows[i]
        if not raw:
            continue
        cells = [(v or '').strip() for v in raw]
        if not any(cells):
            continue

        a = cell(raw, col_map, 'name') or (cells[0] if cells else '')
        b = cell(raw, col_map, 'link') or (cells[1] if len(cells) > 1 else '')

        if not a:
            continue
        # Section row?
        if is_section_row(cells):
            cur_section = a
            stats['section_rows'] += 1
            continue
        if b.lower() == 'duplicate record':
            stats['skipped_dup_marker'] += 1
            continue
        if len(a) > 150:
            stats['skipped_junk'] += 1
            continue

        rec = {
            'name': a, 'section': cur_section, 'link': b,
            'contact': cell(raw, col_map, 'contact'),
            'physical': cell(raw, col_map, 'physical'),
            'mental': cell(raw, col_map, 'mental'),
            'spiritual': cell(raw, col_map, 'spiritual'),
            'emotional': cell(raw, col_map, 'emotional'),
            'survivors': cell(raw, col_map, 'survivors'),
            'youth': cell(raw, col_map, 'youth'),
            'connect': cell(raw, col_map, 'connect'),
            'population': cell(raw, col_map, 'population'),
            'populationNote': cell(raw, col_map, 'populationNote'),
            'additionalLinks': cell(raw, col_map, 'additionalLinks'),
            'strategicPlan': cell(raw, col_map, 'strategicPlan'),
            'agm': cell(raw, col_map, 'agm'),
            'financials': cell(raw, col_map, 'financials'),
            'notes': cell(raw, col_map, 'notes'),
            'direction': cell(raw, col_map, 'direction') or None,
            'type': cell(raw, col_map, 'type') or None,
            'sourceRow': i + 1,
        }

        # Optional explicit lat/lon
        for k in ('lat', 'lon'):
            v = cell(raw, col_map, k)
            if v:
                try:
                    rec[k] = float(v)
                except ValueError:
                    pass

        classify(rec)
        rec['staff'] = parse_contacts(rec['contact'])

        if not rec['contact']:
            stats['no_contact'] += 1
        if not any(rec[k] for k in ('physical', 'mental', 'spiritual',
                                    'emotional', 'youth', 'survivors')):
            stats['no_narrative'] += 1

        # Preserve lat/lon from previous if not explicit
        if previous and not rec.get('lat'):
            nk = normalize(rec['name'])
            for prev in previous:
                if normalize(prev.get('name', '')) == nk:
                    if prev.get('lat') is not None:
                        rec['lat'] = prev['lat']
                    if prev.get('lon') is not None:
                        rec['lon'] = prev['lon']
                    break

        # Dedupe: merge non-empty fields onto existing record
        k = normalize(rec['name'])
        if k in dedup:
            ex = records[dedup[k]]
            for field, val in rec.items():
                if not ex.get(field) and val:
                    ex[field] = val
            stats['merged_dupes'] += 1
        else:
            dedup[k] = len(records)
            records.append(rec)

    if verbose:
        print(f'[parse] {len(records)} unique records', file=sys.stderr)
        print(f'[parse] {stats["section_rows"]} section rows skipped, '
              f'{stats["skipped_dup_marker"]} dup-marker rows, '
              f'{stats["skipped_junk"]} junk rows, '
              f'{stats["merged_dupes"]} dupes merged', file=sys.stderr)
        print(f'[quality] {stats["no_contact"]} records have no contact info, '
              f'{stats["no_narrative"]} have no narrative content', file=sys.stderr)

    return records


def load_previous_dataset(js_path: Path) -> list[dict]:
    """Read an existing communities-data.js to lift forward lat/lon."""
    if not js_path.exists():
        return []
    text = js_path.read_text(encoding='utf-8')
    m = re.search(r'window\.COMMUNITIES\s*=\s*(\[.*?\])\s*;?\s*$', text, re.S)
    if not m:
        return []
    try:
        return json.loads(m.group(1))
    except json.JSONDecodeError:
        return []


def write_dataset(records: list[dict], out_path: Path) -> None:
    payload = json.dumps(records, ensure_ascii=False, indent=2)
    out_path.write_text(
        f'// Auto-generated by tools/process_sheet.py — do not edit by hand.\n'
        f'window.COMMUNITIES = {payload};\n',
        encoding='utf-8',
    )


def main():
    ap = argparse.ArgumentParser(
        description='Process the Mino Bimaadiziwin master sheet (Excel or CSV)')
    ap.add_argument('input', help='Path to the master .xlsx, .xls or .csv file')
    ap.add_argument('--out', default='communities-data.js',
                    help='Output JS file (default: communities-data.js)')
    ap.add_argument('--sheet', default=None,
                    help='Specific worksheet name (Excel only). '
                         'Default: "Master Sheet" or the first sheet.')
    ap.add_argument('--quiet', action='store_true', help='Suppress diagnostics')
    args = ap.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.out)
    verbose = not args.quiet

    if not in_path.exists():
        sys.exit(f'File not found: {in_path}')

    ext = in_path.suffix.lower()
    if ext in ('.xlsx', '.xls'):
        rows, src = load_xlsx_rows(in_path, args.sheet)
        if verbose:
            print(f'[input] {len(rows)} rows from sheet "{src}"', file=sys.stderr)
    elif ext == '.csv':
        rows, src = load_csv_rows(in_path)
        if verbose:
            print(f'[input] {len(rows)} rows from {src}', file=sys.stderr)
    else:
        sys.exit(f'Unsupported format: {ext}. Use .xlsx, .xls, or .csv')

    previous = load_previous_dataset(out_path)
    if previous and verbose:
        print(f'[coords] carrying forward lat/lon from {len(previous)} existing records',
              file=sys.stderr)

    records = process_rows(rows, previous, verbose=verbose)
    write_dataset(records, out_path)
    if verbose:
        print(f'\u2713 wrote {len(records)} records → {out_path}', file=sys.stderr)


if __name__ == '__main__':
    main()
