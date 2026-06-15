"""State-of-the-art multi-sheet workbook ingestion.

The master `.xlsx` has SEVEN sheets that we cross-reference:

  - 'Master Sheet'           — canonical narrative data (one row per community)
  - 'List of 85 communities' — the official list of supported communities AND
                               a second column listing communities found in the
                               Master Sheet that are NOT in the official 85
  - 'Missing Information'    — per-community, per-field WHAT is missing + why
  - 'Needs Review'           — per-community, per-field WHAT needs review
  - 'Correction sheet'       — corrections that have been applied / are queued
  - 'Additional links'       — extra organisation links (umbrella orgs)
  - 'Colour Codes'           — legend (cosmetic, skipped)

This module reads them ALL and produces:

  - records: enriched community records (same shape as tools/process_sheet.py)
  - coverage85: per-canonical-community status (in master? notes? matched-to?)
  - annotations: per-community-per-field missing/review/correction items
  - workbookMeta: which sheets we found, how many rows each, etc.

If a workbook only has a single sheet (the legacy format), everything still
works — the cross-reference panels just come back empty.
"""
from __future__ import annotations
import difflib
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Reuse the canonical-field detector and contact parser from process_sheet.
from . import processor  # for processor._load_sheet_module()


def _ps():
    """Lazy import the existing process_sheet helpers."""
    return processor._load_sheet_module()


# --- normalisation -------------------------------------------------------- #
#
# Community name matching is fuzzy. Real-world examples we need to match:
#
#   Master sheet                                           ↔  85-list
#   ─────────────────────────────────────────────────────  ↔  ────────────────────
#   Batchewana First Nation (Association of Iroquois...)   ↔  Batchewana First Nation
#   Big Trout Lake #272 / Kitchenuhmaykoosib Inninuwug FN  ↔  Big Trout Lake #272 / Kitchenuhmaykoosib...
#   Cree Nation of Mistissini                              ↔  The Mistassini/Cree First Nation  (renamed)
#   Mattagami First Nation (Nishnawbe Aski Nation - Wa...) ↔  Mattagami First Nation
#
# Strategy: every name produces a SET of candidate keys. Match if any overlap.
# Keys ignore: parenthetical suffixes, "#NNN" band numbers, slash-separated
# alternate spellings, "The " prefix, whitespace, punctuation, and the
# "First Nation"/"First Nations" suffix (normalized to "fn").

_NORM_RX = re.compile(r"[\s\(\)\/,'\-\.#&]")
_FN_RX = re.compile(r"firstnations?")
_HASH_NUM_RX = re.compile(r"#\s*\d+")
_PAREN_RX = re.compile(r"\s*\([^)]*\)")
_THE_PREFIX_RX = re.compile(r"^the\s+")

# Region keywords that appear in the right-side "extras" column of the 85-list
# sheet — used to extract a region hint and drop it from the name itself.
_REGION_HINTS = {"north", "south", "east", "west", "central",
                 "algoma district first nation communities"}


def _normalize_key(fragment: str) -> str:
    """Reduce a name fragment to a comparable key."""
    s = unicodedata.normalize("NFKD", str(fragment)).encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = _THE_PREFIX_RX.sub("", s)
    s = _HASH_NUM_RX.sub("", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = _NORM_RX.sub("", s)
    s = _FN_RX.sub("fn", s)
    return s.strip()


def name_keys(name: str) -> set[str]:
    """Return all candidate keys for a community name (handles aliases)."""
    if not name:
        return set()
    s = unicodedata.normalize("NFKD", str(name)).encode("ascii", "ignore").decode("ascii")
    # Drop parenthetical suffixes like "(AIAI)" or "(Nishnawbe Aski Nation - …)"
    s = _PAREN_RX.sub("", s)
    s = re.sub(r"\s+", " ", s).strip()
    # Slash-separated alternates: each side is a candidate
    parts = [p.strip() for p in s.split("/") if p.strip()]
    keys: set[str] = set()
    for p in parts:
        k = _normalize_key(p)
        if k:
            keys.add(k)
    # Also store the full string as a key (some names contain `/` in their official form)
    full = _normalize_key(s)
    if full:
        keys.add(full)
    return keys


# Tokens too generic to base a match on alone.
_GENERIC_TOKENS = {
    "first", "nation", "nations", "fn", "the", "of", "and", "council",
    "tribe", "tribal", "community", "communities", "band", "people",
    "peoples", "association", "centre", "center", "organization",
    "indigenous", "anishinaabe", "anishnaabe", "anishnawbek", "mi'kmaq",
    "cree", "ojibwe", "ojibway", "river", "lake", "bay", "island",
    "the", "st", "st.", "saint",
}


def _tokens(name: str) -> set[str]:
    """Significant tokens from a name — for token-overlap matching."""
    if not name:
        return set()
    s = unicodedata.normalize("NFKD", str(name)).encode("ascii", "ignore").decode("ascii")
    s = _PAREN_RX.sub("", s)
    s = _HASH_NUM_RX.sub("", s)
    s = s.lower()
    toks = re.findall(r"[a-z']+", s)
    sig = {t for t in toks if t not in _GENERIC_TOKENS and len(t) >= 3}
    return sig


def _fuzzy_ratio(a: str, b: str) -> float:
    """0..1 string similarity via difflib (handles typos like Mistassini/Mistissini)."""
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def smart_match(name: str, candidates: list[dict], name_field: str = "name") -> dict | None:
    """Find the best candidate match for `name` from `candidates`.

    Four-pass strategy (most-precise first):

      1. EXACT       — name_keys() overlap (handles parens, slashes, hash numbers)
      2. WEIGHTED    — shared tokens scored by uniqueness × length. A token that
                       appears in only ONE candidate (e.g., "Akwesasne") is much
                       stronger evidence than one in many (e.g., "Chippewas").
      3. FUZZY-TOK   — at least one pair of distinct long tokens has >= 0.85 fuzzy
                       similarity AND ratio < 1.0 (catches typos like
                       "Mistassini" ↔ "Mistissini"; rejects common-word false
                       positives where ratio == 1.0).
      4. FUZZY-ALL   — full normalised string similarity >= 0.90 (last resort).
    """
    if not name or not candidates:
        return None

    target_keys = name_keys(name)
    target_tokens = _tokens(name)

    # Pass 1 — EXACT key overlap.
    for c in candidates:
        c_keys = c.get("_keys") or name_keys(c.get(name_field, ""))
        if target_keys & c_keys:
            return c

    # Pass 2 — WEIGHTED token overlap.
    # Build token frequency across the candidate pool — rare tokens are gold.
    from collections import Counter
    token_freq: Counter = Counter()
    for c in candidates:
        c_tokens = c.get("_tokens") or _tokens(c.get(name_field, ""))
        for t in c_tokens:
            token_freq[t] += 1
    best = None
    best_score = 0.0
    for c in candidates:
        c_tokens = c.get("_tokens") or _tokens(c.get(name_field, ""))
        shared = target_tokens & c_tokens
        if not shared:
            continue
        score = 0.0
        for t in shared:
            uniqueness = 1.0 / token_freq[t]  # 1.0 if unique, 0.25 if 4 candidates have it
            score += uniqueness * len(t)
        if score > best_score:
            best_score = score
            best = c
    # Threshold 6.0 means: ONE 6+ char token unique to 1 candidate (1.0×6=6),
    # or one 9+ char token in <=2 candidates (e.g., 0.5×9=4.5 would FAIL — good,
    # that's the "Chippewas of Sarnia/Nawash" common-word case). Multiple shared
    # tokens easily clear the bar.
    if best and best_score >= 6.0:
        return best

    # Pass 3 — fuzzy token match (catches single-letter typos).
    # Skips:
    #   - identical tokens (already considered in Pass 2)
    #   - tokens where one is a substring of the other (e.g., "chippewa"/
    #     "chippewas" is stem variation, not a typo — those should rely on
    #     other distinctive tokens to match in Pass 2 instead).
    best = None
    best_ratio = 0.0
    for c in candidates:
        c_tokens = c.get("_tokens") or _tokens(c.get(name_field, ""))
        for tt in target_tokens:
            if len(tt) < 6:
                continue
            for ct in c_tokens:
                if len(ct) < 6 or tt == ct:
                    continue
                if tt in ct or ct in tt:
                    continue  # substring relation → not a typo
                r = _fuzzy_ratio(tt, ct)
                if 0.85 <= r < 1.0 and r > best_ratio:
                    best_ratio = r
                    best = c
    if best:
        return best

    # Pass 4 — full-string fuzzy (very strict threshold).
    target_full = max(target_keys, key=len) if target_keys else ""
    best = None
    best_ratio = 0.0
    for c in candidates:
        c_keys = c.get("_keys") or name_keys(c.get(name_field, ""))
        for k in c_keys:
            r = _fuzzy_ratio(target_full, k)
            if r > best_ratio:
                best_ratio = r
                best = c
    if best and best_ratio >= 0.90:
        return best

    return None


def normalize(name: str) -> str:
    """Primary normalized key — kept for backward compat with the rest of the code.

    Returns the first key from name_keys(), or empty string. New code should
    use name_keys() and check for set intersection.
    """
    keys = name_keys(name)
    return next(iter(sorted(keys, key=len, reverse=True))) if keys else ""


# Pre-clean fragments that show up inside parens in the 85-list right column
PAREN_REGION_RX = re.compile(r"\s*\(\s*([^)]+)\s*\)\s*")


def clean_community_name(name: str) -> tuple[str, str]:
    """Return (clean_name, region_hint). Strips trailing `(North)`/region tags."""
    if not name:
        return "", ""
    s = re.sub(r"\s+", " ", str(name)).strip()
    m = PAREN_REGION_RX.search(s)
    region = ""
    if m:
        candidate = m.group(1).strip()
        # Only treat the parens as a region hint if it's actually a region.
        if candidate.lower() in _REGION_HINTS:
            region = candidate
            s = PAREN_REGION_RX.sub("", s).strip()
    return s, region


# --- workbook readers ----------------------------------------------------- #

def _read_sheet(wb, name: str) -> list[list[str]]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = []
    for raw in ws.iter_rows(values_only=True):
        if not raw:
            continue
        rows.append(["" if v is None else str(v) for v in raw])
    return rows


# Header signatures for each known sheet type. score_sheet_as() scans the
# first 5 rows of any sheet, counts how many of these patterns it sees,
# and picks the highest-scoring sheet for each type. This makes detection
# robust to sheet renames, reordering, garbage tabs, and the original
# 'Master Sheet' being moved or duplicated.

_MASTER_SIGNATURES = [
    r"community.*link", r"contact.*info", r"physical.*health",
    r"mental.*health", r"spiritual", r"emotional",
    r"survivor", r"youth", r"connect.*survivors?.*youth",
    r"population", r"strategic.*plan", r"annual.*general", r"financial",
]

_LIST85_SIGNATURES = [
    r"list.*85.*communit", r"\b85\s+communit", r"\bfirst nation\b",
    r"\bnumber\b", r"\bcommunity\b",
]

_MISSING_SIGNATURES = [
    r"community/organi[sz]ation", r"specific.*item", r"\bstatus\b",
    r"\bdi?escription\b", r"missing.*information",
]

_REVIEW_SIGNATURES = [
    r"community/organi[sz]ation", r"specific.*item",
    r"needs.*review", r"\bstatus\b",
]

_CORRECTION_SIGNATURES = [
    r"community/organi[sz]ation", r"specific.*item",
    r"\bdi?escription\b", r"correction",
]

_ADDITIONAL_SIGNATURES = [
    r"additional.*link", r"community.*link", r"physical.*health",
]


def _score_sheet_against(sheet_name: str, rows: list[list[str]],
                         signatures: list[str]) -> int:
    """Return a 0..N score for how strongly this sheet matches a signature list.

    Looks at:
      - the sheet name itself (gentle bonus if any signature pattern matches)
      - the first 5 rows of cell text (main signal — these are usually headers)
    """
    score = 0
    name_blob = sheet_name.lower()
    cell_blob = ""
    for r in rows[:5]:
        for c in r:
            if c:
                cell_blob += " " + str(c).lower()
    for pat in signatures:
        if re.search(pat, name_blob):
            score += 1
        if re.search(pat, cell_blob):
            score += 2
    # Bonus for sheets with substantial data (real master sheet has many rows).
    nonempty = sum(1 for r in rows if any(c and str(c).strip() for c in r))
    if nonempty >= 50:
        score += 3
    elif nonempty >= 20:
        score += 1
    return score


def detect_sheets(wb) -> dict[str, str | None]:
    """Auto-detect which sheet plays each role, by header content (not name).

    Returns a dict like:
      { 'master': 'Master Sheet',
        'list85': 'List of 85 communities',
        'missing': 'Missing Information',
        'review':  'Needs Review',
        'correction': 'Correction sheet',
        'additional': 'Additional links' }

    Each value can be None if no sheet scored high enough. Garbage tabs
    (Colour Codes, Notes, Sandbox, etc.) end up unassigned.
    """
    # Pre-load row data for every sheet once.
    sheets_data = [(name, _read_sheet(wb, name)) for name in wb.sheetnames]

    def best(signatures: list[str], min_score: int = 5,
             exclude: set[str] | None = None) -> str | None:
        exclude = exclude or set()
        ranked = []
        for name, rows in sheets_data:
            if name in exclude or not rows:
                continue
            s = _score_sheet_against(name, rows, signatures)
            if s >= min_score:
                ranked.append((s, name))
        ranked.sort(reverse=True)
        return ranked[0][1] if ranked else None

    # Master first — it's the most distinctive (many narrative columns).
    master = best(_MASTER_SIGNATURES, min_score=8)
    used = {master} if master else set()

    list85 = best(_LIST85_SIGNATURES, min_score=4, exclude=used)
    if list85:
        used.add(list85)

    # Missing / Review / Correction all share most columns — disambiguate by
    # the distinguishing keyword in name + headers.
    missing = best(_MISSING_SIGNATURES, min_score=5, exclude=used)
    if missing:
        used.add(missing)
    review = best(_REVIEW_SIGNATURES, min_score=5, exclude=used)
    if review:
        used.add(review)
    correction = best(_CORRECTION_SIGNATURES, min_score=5, exclude=used)
    if correction:
        used.add(correction)

    additional = best(_ADDITIONAL_SIGNATURES, min_score=4, exclude=used)

    return {
        "master": master,
        "list85": list85,
        "missing": missing,
        "review": review,
        "correction": correction,
        "additional": additional,
    }


def _find_sheet(wb, candidates: list[str]) -> str | None:
    """Find the first sheet whose name matches any candidate (case-insensitive)."""
    lowered = {s.lower().strip(): s for s in wb.sheetnames}
    for c in candidates:
        if c.lower() in lowered:
            return lowered[c.lower()]
    # Fuzzy: contains
    for c in candidates:
        for low, real in lowered.items():
            if c.lower() in low:
                return real
    return None


# --- structured readers per sheet ---------------------------------------- #

# Hyperlink extraction is done by reading the .xlsx zip's XML directly rather
# than re-opening the workbook with openpyxl in non-read-only mode. The latter
# materialises every cell (styles, formats) for the WHOLE workbook and, done
# once per sheet, spiked memory enough to crash a small instance on upload.
# The zip/XML path below parses two tiny XML files per sheet and uses almost
# no memory.
_NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_NS_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_NS_PKG = "{http://schemas.openxmlformats.org/package/2006/relationships}"


def _ref_to_rc(ref: str) -> tuple[int, int]:
    """'B3' → (row=3, col_index=1, 0-based). Returns (0,-1) if unparseable."""
    letters = "".join(ch for ch in ref if ch.isalpha())
    digits = "".join(ch for ch in ref if ch.isdigit())
    if not letters or not digits:
        return (0, -1)
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch.upper()) - 64)
    return (int(digits), col - 1)


def _extract_hyperlinks(path) -> dict:
    """Low-memory hyperlink reader straight from the .xlsx zip.

    Returns {sheet_name: {row_number: {col_index: url}}} for EXTERNAL http(s)
    links only. Internal workbook anchors (location=…, no r:id) are skipped.
    Row numbers are real 1-based spreadsheet rows, which equal each record's
    sourceRow (process_rows numbers rows 1:1 with the sheet).
    """
    import zipfile
    import xml.etree.ElementTree as ET

    out: dict = {}
    if not path:
        return out
    try:
        zf = zipfile.ZipFile(str(path))
    except Exception:
        return out
    try:
        names = set(zf.namelist())
        wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        name_to_rid = {}
        for sh in wb_xml.iter(_NS_MAIN + "sheet"):
            nm, rid = sh.get("name"), sh.get(_NS_REL + "id")
            if nm and rid:
                name_to_rid[nm] = rid
        wb_rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_to_path = {}
        for rel in wb_rels.iter(_NS_PKG + "Relationship"):
            tgt = rel.get("Target") or ""
            wpath = tgt if tgt.startswith("xl/") else "xl/" + tgt.lstrip("/")
            rid_to_path[rel.get("Id")] = wpath

        for sheet_name, rid in name_to_rid.items():
            wpath = rid_to_path.get(rid)
            if not wpath or wpath not in names:
                continue
            try:
                sx = ET.fromstring(zf.read(wpath))
            except Exception:
                continue
            hlinks = sx.find(_NS_MAIN + "hyperlinks")
            if hlinks is None:
                continue
            base = wpath.rsplit("/", 1)[-1]
            rels_path = wpath.rsplit("/", 1)[0] + "/_rels/" + base + ".rels"
            rid_target = {}
            if rels_path in names:
                try:
                    sr = ET.fromstring(zf.read(rels_path))
                    for rel in sr.iter(_NS_PKG + "Relationship"):
                        rid_target[rel.get("Id")] = rel.get("Target")
                except Exception:
                    pass
            grid: dict = {}
            for h in hlinks.findall(_NS_MAIN + "hyperlink"):
                ref = h.get("ref")
                hid = h.get(_NS_REL + "id")
                if not ref or not hid:
                    continue  # internal location-only link → skip
                url = rid_target.get(hid)
                if not url or not str(url).lower().startswith(("http://", "https://")):
                    continue
                row, col = _ref_to_rc(ref.split(":")[0])
                if row <= 0 or col < 0:
                    continue
                grid.setdefault(row, {})[col] = str(url).strip()
            if grid:
                out[sheet_name] = grid
    except Exception:
        return out
    finally:
        try:
            zf.close()
        except Exception:
            pass
    return out


# Canonical-field → which record key the front-end reads it under.
_FIELD_LINK_ALIAS = {
    "link": "website", "contact": "contacts", "additionalLinks": "links",
    "populationNote": "popInfo",
}


def _attach_field_links(records: list[dict], rows: list[list[str]],
                        grid: dict, ps) -> None:
    """Map each row's per-column hyperlinks onto its record as `fieldLinks`,
    keyed by the front-end field name. `grid` is {row_number: {col: url}};
    alignment is by sourceRow (== real spreadsheet row)."""
    if not grid:
        return
    hdr_idx, col_map = ps.detect_header(rows)
    idx_to_field = {ci: field for field, ci in col_map.items()}
    for rec in records:
        sr = rec.get("sourceRow")
        row_links = grid.get(sr)
        if not row_links:
            continue
        field_links: dict[str, str] = {}
        for ci, url in row_links.items():
            field = idx_to_field.get(ci)
            if not field or field in ("lat", "lon", "type", "direction", "population"):
                continue
            key = _FIELD_LINK_ALIAS.get(field, field)
            field_links.setdefault(key, url)
        if field_links:
            rec["fieldLinks"] = field_links


def read_master_sheet(wb, sheet: str | None = None, link_grid=None) -> list[dict]:
    """Master Sheet → enriched community records, via tools/process_sheet.py.

    `sheet` can be passed explicitly (preferred — comes from content-based
    detection in ingest_workbook). If not, fall back to name-based search.
    `link_grid` ({row: {col: url}}, pre-extracted once in ingest_workbook)
    lets the source link behind each cell survive into the dashboard.
    """
    if not sheet:
        sheet = _find_sheet(wb, ["Master Sheet", "Master", "Programs"])
    if not sheet:
        sheet = wb.sheetnames[0]
    rows = _read_sheet(wb, sheet)
    if not rows:
        return []
    ps = _ps()
    records = ps.process_rows(rows, previous=None, verbose=False)
    try:
        _attach_field_links(records, rows, link_grid or {}, ps)
    except Exception:
        pass  # links are a bonus — never let them break ingestion
    return records


def read_85_list(wb, sheet: str | None = None) -> dict:
    """Returns {'official': [{number, name, note}], 'inMasterNotInList': [{number, name, region}]}."""
    if not sheet:
        sheet = _find_sheet(wb, ["List of 85 communities", "List of communities", "85 communities"])
    if not sheet:
        return {"official": [], "inMasterNotInList": []}
    rows = _read_sheet(wb, sheet)

    official, extra = [], []
    # Header is in row 1 ("List of the 85 communities") with sub-header in row 2
    # ("Number" / "First Nation" on the left, "Number" / "Community" on the right).
    for r in rows[2:]:  # skip both header rows
        # Left side: official 85 list (cols 0-2: Number, FirstNation, Notes)
        number = r[0] if len(r) > 0 else ""
        fn = r[1] if len(r) > 1 else ""
        note = r[2] if len(r) > 2 else ""
        if fn and not fn.lower().startswith("number"):
            clean, region = clean_community_name(fn)
            try:
                num = int(float(number)) if number else None
            except ValueError:
                num = None
            official.append({
                "number": num,
                "name": clean,
                "note": (note or "").strip() or None,
                "region": region or None,
                "key": normalize(clean),
            })
        # Right side: in-master-not-in-list (cols 4-5)
        en = r[4] if len(r) > 4 else ""
        ef = r[5] if len(r) > 5 else ""
        if ef and not ef.lower().startswith("number") and not ef.lower().startswith("community"):
            clean, region = clean_community_name(ef)
            try:
                num = int(float(en)) if en else None
            except ValueError:
                num = None
            extra.append({
                "number": num,
                "name": clean,
                "region": region or None,
                "key": normalize(clean),
            })

    return {"official": official, "inMasterNotInList": extra}


def read_annotations(wb, sheet_name: str, kind: str, sheet: str | None = None,
                     link_grid=None) -> list[dict]:
    """Generic reader for Missing Information / Needs Review / Correction sheet.

    When `link_grid` ({row: {col: url}}) is given, the EXTERNAL hyperlinks in
    each row's cells (e.g. the "Consolidated Financial Statements March 31,
    2023" link inside a Needs-Review note) are attached to the entry as `links`,
    so the COMING SOON / UNDER REVIEW cards can show them clickably.
    """
    if not sheet:
        sheet = _find_sheet(wb, [sheet_name])
    if not sheet:
        return []
    rows = _read_sheet(wb, sheet)
    link_grid = link_grid or {}
    out = []
    for i, r in enumerate(rows[1:], start=1):  # skip header row
        community = r[0] if len(r) > 0 else ""
        item = r[1] if len(r) > 1 else ""
        if not community and not item:
            continue
        # Correction sheet header is slightly different (no Status column).
        if kind == "correction":
            entry = {
                "kind": kind,
                "community": (community or "").strip(),
                "communityKey": normalize(community),
                "item": (item or "").strip(),
                "description": (r[2] if len(r) > 2 else "").strip(),
                "date": (r[3] if len(r) > 3 else "").strip(),
                "note": (r[4] if len(r) > 4 else "").strip(),
            }
        else:
            entry = {
                "kind": kind,
                "community": (community or "").strip(),
                "communityKey": normalize(community),
                "item": (item or "").strip(),
                "status": (r[2] if len(r) > 2 else "").strip(),
                "description": (r[3] if len(r) > 3 else "").strip(),
                "date": (r[4] if len(r) > 4 else "").strip(),
                "note": (r[5] if len(r) > 5 else "").strip(),
            }
        # Attach any external hyperlinks found in this row (skip the community
        # name cell in col 0, which often links to a generic homepage). The data
        # row at list-index i is real spreadsheet row i+1.
        row_links = link_grid.get(i + 1)
        if row_links:
            seen, uniq = set(), []
            for ci, u in sorted(row_links.items()):
                if ci >= 1 and u not in seen:
                    seen.add(u); uniq.append(u)
            if uniq:
                entry["links"] = uniq
        out.append(entry)
    return out


def read_additional_links(wb, sheet: str | None = None) -> list[dict]:
    if not sheet:
        sheet = _find_sheet(wb, ["Additional links", "Additional Links"])
    if not sheet:
        return []
    rows = _read_sheet(wb, sheet)
    out = []
    cur_section = ""
    for r in rows[1:]:
        a = r[0] if len(r) > 0 else ""
        b = r[1] if len(r) > 1 else ""
        if not a:
            continue
        # Section row (single col)
        if a and not b:
            cur_section = a.strip()
            continue
        out.append({
            "name": a.strip(),
            "link": b.strip(),
            "section": cur_section,
            "key": normalize(a),
        })
    return out


# --- main entry ----------------------------------------------------------- #

def ingest_workbook(path: Path) -> dict:
    """Read the whole workbook and produce a structured snapshot.

    Detection of which sheet plays which role is CONTENT-BASED, not name-
    based — see detect_sheets(). Garbage tabs ('Colour Codes', 'Notes',
    'Sandbox') are ignored. A renamed Master Sheet still works.
    """
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    detected = detect_sheets(wb)

    # Extract every sheet's external hyperlinks ONCE, cheaply, from the zip XML
    # (no second openpyxl load — that previously spiked memory on upload).
    hyperlinks = {}
    try:
        hyperlinks = _extract_hyperlinks(path)
    except Exception:
        hyperlinks = {}

    master_records = read_master_sheet(wb, detected["master"],
                                       link_grid=hyperlinks.get(detected["master"]))
    list85 = read_85_list(wb, detected["list85"])
    missing = read_annotations(wb, "Missing Information", "missing", detected["missing"],
                               link_grid=hyperlinks.get(detected["missing"]))
    review = read_annotations(wb, "Needs Review", "review", detected["review"],
                              link_grid=hyperlinks.get(detected["review"]))
    corrections = read_annotations(wb, "Correction sheet", "correction", detected["correction"],
                                   link_grid=hyperlinks.get(detected["correction"]))
    extra_links = read_additional_links(wb, detected["additional"])

    # Precompute keys and tokens for every entry (one-time, makes smart_match fast).
    for rec in master_records:
        rec["_keys"] = name_keys(rec.get("name", ""))
        rec["_tokens"] = _tokens(rec.get("name", ""))
    for entry in list85["official"]:
        entry["_keys"] = name_keys(entry["name"])
        entry["_tokens"] = _tokens(entry["name"])
    for entry in list85["inMasterNotInList"]:
        entry["_keys"] = name_keys(entry["name"])
        entry["_tokens"] = _tokens(entry["name"])
    all_annotations = []
    for batch in (missing, review, corrections):
        for a in batch:
            a["_keys"] = name_keys(a["community"])
            a["_tokens"] = _tokens(a["community"])
            all_annotations.append(a)

    # Cross-reference each master record against the 85 list + extras + annotations.
    # Uses the 3-pass smart matcher (exact → token-overlap → fuzzy).
    for rec in master_records:
        rec["nameKey"] = sorted(rec["_keys"])[0] if rec["_keys"] else ""

        list85_match = smart_match(rec.get("name", ""), list85["official"])
        extra_match = (smart_match(rec.get("name", ""), list85["inMasterNotInList"])
                       if not list85_match else None)

        rec["in85List"] = list85_match is not None
        rec["in85Extra"] = extra_match is not None
        if list85_match:
            rec["officialNumber"] = list85_match["number"]
            if list85_match.get("note"):
                rec["listNote"] = list85_match["note"]
        elif extra_match:
            rec["officialNumber"] = None
            rec["extraSection"] = extra_match.get("region")

        # Annotations: match each annotation entry against this record, dedupe.
        seen = set()
        record_anns: list[dict] = []
        for a in all_annotations:
            # An annotation belongs to this record if its name maps to it.
            if rec["_keys"] & a["_keys"] or (
                len(rec["_tokens"] & a["_tokens"]) >= 2
            ):
                sig = (a.get("kind"), a.get("item"), a.get("date"), a.get("description", "")[:50])
                if sig in seen:
                    continue
                seen.add(sig)
                record_anns.append(a)
        rec["annotations"] = record_anns
        rec["missingCount"] = sum(1 for a in record_anns if a["kind"] == "missing")
        rec["reviewCount"] = sum(1 for a in record_anns if a["kind"] == "review")
        rec["correctionCount"] = sum(1 for a in record_anns if a["kind"] == "correction")

    # Coverage85: for every official community, was a master sheet record found?
    coverage_items = []
    for entry in list85["official"]:
        matched = smart_match(entry["name"], master_records)
        coverage_items.append({
            "number": entry["number"],
            "name": entry["name"],
            "note": entry["note"],
            "inMaster": matched is not None,
            "matchedName": matched["name"] if matched else None,
            "matchedDirection": (matched or {}).get("direction"),
            "matchedAnnotations": len((matched or {}).get("annotations", [])),
        })

    # Strip internal _keys/_tokens before serialising (they're sets, not JSON-friendly).
    def _clean(d: dict) -> dict:
        return {k: v for k, v in d.items() if not k.startswith("_")}
    master_records = [_clean(r) for r in master_records]
    list85["official"] = [_clean(e) for e in list85["official"]]
    list85["inMasterNotInList"] = [_clean(e) for e in list85["inMasterNotInList"]]
    for batch in (missing, review, corrections):
        for a in batch:
            a.pop("_keys", None)
            a.pop("_tokens", None)

    return {
        "records": master_records,
        "list85": list85,
        "coverage85": coverage_items,
        "annotations": {
            "missing": missing,
            "review": review,
            "corrections": corrections,
            "totals": {
                "missing": len(missing),
                "review": len(review),
                "corrections": len(corrections),
            },
        },
        "additionalLinks": extra_links,
        "meta": {
            "sheets": wb.sheetnames,
            "detectedSheets": detected,
            "masterRows": len(master_records),
            "list85Official": len(list85["official"]),
            "list85Extra": len(list85["inMasterNotInList"]),
            "unusedSheets": [s for s in wb.sheetnames if s not in detected.values()],
        },
    }
