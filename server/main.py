"""Mino Bimaadiziwin Atlas — backend server.

Run from the repo root:

    pip install -r server/requirements.txt
    python -m server.main            # or: uvicorn server.main:app --reload

Endpoints
=========
GET  /api/health
POST /api/auth/login              (form: username, password)
POST /api/auth/logout
GET  /api/auth/me
GET  /api/users                   (admin)
POST /api/users                   (admin) — body: {username, password, role}
DELETE /api/users/{username}      (admin)
POST /api/users/{username}/password  (admin) — body: {password}

GET  /api/communities             — current dataset (records[])
POST /api/communities/upload      (admin) — multipart file
POST /api/communities/{id}/edit   (admin) — body: partial record
DELETE /api/communities/{id}      (admin)
GET  /api/communities/edits       — merged edits
GET  /api/communities/versions    — dataset version history
POST /api/communities/versions/{v}/activate (admin)

GET  /api/analytics/overview
GET  /api/analytics/pillars
GET  /api/analytics/gaps
GET  /api/analytics/population
GET  /api/analytics/keywords
GET  /api/analytics/clusters?k=5
GET  /api/analytics/coverage
GET  /api/analytics/quality
GET  /api/analytics/full

Plus static file serving for the dashboard at `/` and the CMS at `/cms`.
"""
from __future__ import annotations
import os
import shutil
import sys
import time
import uuid
from pathlib import Path

from fastapi import (
    FastAPI, HTTPException, UploadFile, File, Form, Depends, Request, Response, status
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from . import analytics, auth, config, coords, db, enrich, events, gsheets, mailer, processor, workbook


app = FastAPI(
    title="Mino Bimaadiziwin Community Atlas API",
    version="1.0.0",
    description="Backend for the Community Services Atlas dashboard.",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=config.CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def _no_cache_source(request, call_next):
    """Force browsers to revalidate source assets (HTML/JSX/JS/CSS) on every
    load. These files change on every deploy but the CMS/dashboard reference
    several of them without cache-busting query strings, so without this the
    browser serves stale copies and edits appear to "do nothing". Binary/media
    responses (which set their own Cache-Control) are left untouched."""
    response = await call_next(request)
    path = request.url.path.lower()
    if path.endswith(".html") or path == "/" or path.endswith("/"):
        # The HTML is the entry point that references every versioned asset.
        # Never cache it, so a deploy is picked up immediately (no hard-refresh).
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    elif path.endswith((".jsx", ".js", ".css")):
        # don't clobber an explicit cache policy a route already set
        if "cache-control" not in (k.lower() for k in response.headers.keys()):
            response.headers["Cache-Control"] = "no-cache, must-revalidate"
    return response


# ----------------------------- bootstrap ---------------------------------- #

# Set when the database cannot be reached at boot. The site still serves the
# dashboard from the bundled dataset in that case — a dead/paused database must
# never take the whole public site down.
DB_OFFLINE: dict = {"offline": False, "error": ""}


@app.on_event("startup")
def _startup() -> None:
    try:
        db.init_db()
        # load the saved Groq API key (if any) so the AI layer is ready
        try:
            from . import llm as _llm0
            _llm0.set_key(db.get_settings().get("ai.groq_key", ""))
        except Exception:
            pass
        # Seed dataset from communities-data.js if DB has none yet.
        # ALSO: when a deploy ships a NEWER bundled dataset (the repo's
        # communities-data.js changed), re-seed it as a new dataset version so the
        # live site reflects the latest sheet without a manual upload. Previous
        # versions (including manual uploads) stay in history and can be
        # re-activated from the CMS at any time.
        import hashlib
        try:
            _bundle_fp = hashlib.sha256(
                config.COMMUNITIES_JS.read_bytes()).hexdigest() if config.COMMUNITIES_JS.exists() else ""
        except Exception:
            _bundle_fp = ""
        if db.current_dataset() is None:
            records = processor.load_initial_records()
            if records:
                db.save_dataset(records, source_filename="communities-data.js (seed)",
                                uploaded_by="system")
                db.log_audit("system", "seed", "dataset",
                             f"Seeded {len(records)} records from communities-data.js")
                if _bundle_fp:
                    db.set_setting("seed.fingerprint", _bundle_fp, "system")
        elif _bundle_fp and db.get_settings().get("seed.fingerprint") != _bundle_fp:
            try:
                records = processor.load_initial_records()
                if records:
                    db.save_dataset(records, source_filename="communities-data.js (auto-reseed)",
                                    uploaded_by="system")
                    db.log_audit("system", "seed", "dataset",
                                 f"Re-seeded {len(records)} records — bundled dataset changed on deploy")
                db.set_setting("seed.fingerprint", _bundle_fp, "system")
            except Exception:
                pass
    except Exception as exc:                       # noqa: BLE001 — never die at boot
        DB_OFFLINE["offline"] = True
        DB_OFFLINE["error"] = f"{type(exc).__name__}: {exc}"
        print("=" * 72, flush=True)
        print("DATABASE UNREACHABLE — starting in READ-ONLY FALLBACK mode.", flush=True)
        print("The dashboard will still load using the bundled dataset.", flush=True)
        print("Admin/CMS features stay disabled until the database is back.", flush=True)
        print(f"Reason: {DB_OFFLINE['error'][:400]}", flush=True)
        print("=" * 72, flush=True)


# ----------------------------- schemas ------------------------------------ #

class LoginOut(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict


class UserIn(BaseModel):
    username: str = Field(min_length=2, max_length=80)
    password: str = Field(min_length=4, max_length=200)
    role: str = "admin"
    email: str | None = None


class PasswordIn(BaseModel):
    password: str = Field(min_length=4, max_length=200)


class EmailIn(BaseModel):
    email: str | None = None


class ResetRequestIn(BaseModel):
    username: str


class ResetConfirmIn(BaseModel):
    token: str = Field(min_length=10)
    password: str = Field(min_length=4, max_length=200)


class NavItemIn(BaseModel):
    id: int | None = None
    slot: str = "main"
    position: int = 0
    label: str = Field(min_length=1, max_length=80)
    view: str = Field(min_length=1, max_length=80)
    icon: str | None = None
    visible: bool = True


class LayoutIn(BaseModel):
    blocks: list[dict]


class StyleIn(BaseModel):
    properties: dict = Field(default_factory=dict)
    customCss: str | None = ""


class EditIn(BaseModel):
    fields: dict = Field(default_factory=dict)
    staff: list | None = None
    departments: list | None = None


class SettingsIn(BaseModel):
    values: dict


class PageIn(BaseModel):
    slug: str = Field(min_length=1, max_length=80)
    title: str = Field(min_length=1, max_length=200)
    body: str = ""
    visible: bool = True


# ------------------------------ health ------------------------------------ #

@app.get("/api/health")
def health() -> dict:
    if DB_OFFLINE["offline"]:
        # the database is unreachable — say so plainly instead of raising
        return {
            "status": "degraded",
            "version": app.version,
            "database": "offline",
            "mode": "fallback — the dashboard is served from the bundled dataset",
            "detail": DB_OFFLINE["error"][:400],
            "fix": "Resume/restore the Postgres instance, or update DATABASE_URL, then redeploy.",
        }
    try:
        ds = db.current_dataset()
    except Exception as exc:                       # noqa: BLE001
        DB_OFFLINE["offline"] = True
        DB_OFFLINE["error"] = f"{type(exc).__name__}: {exc}"
        return {"status": "degraded", "database": "offline", "detail": str(exc)[:400]}
    backend = "postgres" if db.IS_POSTGRES else "sqlite"
    return {
        "status": "ok",
        "version": app.version,
        "backend": backend,
        "persistent": backend == "postgres",
        "dataset": {
            "loaded": ds is not None,
            "records": ds["record_count"] if ds else 0,
            "version": ds["version"] if ds else None,
        },
    }


# --------------------------- read-aloud proxy ----------------------------- #
# Extract the readable text of an external document (PDF or web page) so the
# dashboard's "Listen" feature can read a linked strategic plan / annual report
# aloud. Runs server-side to sidestep browser CORS and to do PDF parsing.

_READ_DOC_MAX_BYTES = 12 * 1024 * 1024   # don't download more than 12 MB
_READ_DOC_MAX_CHARS = 24_000             # ~25 min of speech; plenty for a listen
_READ_DOC_MAX_PAGES = 60


def _read_doc_host_ok(host: str) -> bool:
    """Block obvious internal / private targets (basic SSRF guard)."""
    h = (host or "").lower()
    if not h:
        return False
    if h in ("localhost", "127.0.0.1", "0.0.0.0", "::1", "metadata", "metadata.google.internal"):
        return False
    if h.startswith(("10.", "192.168.", "169.254.", "172.16.", "172.17.", "172.18.",
                     "172.19.", "172.2", "172.30.", "172.31.")):
        return False
    return True


def _html_to_text(html: str) -> str:
    import re as _re
    # drop script/style/nav/etc, then strip tags and collapse whitespace
    html = _re.sub(r"(?is)<(script|style|noscript|head|nav|footer|svg)[^>]*>.*?</\1>", " ", html)
    html = _re.sub(r"(?is)<br\s*/?>", "\n", html)
    html = _re.sub(r"(?is)</(p|div|li|h[1-6]|tr)>", "\n", html)
    text = _re.sub(r"(?s)<[^>]+>", " ", html)
    # unescape a few common entities
    for a, b in (("&nbsp;", " "), ("&amp;", "&"), ("&quot;", '"'),
                 ("&#39;", "'"), ("&rsquo;", "'"), ("&lsquo;", "'"),
                 ("&ldquo;", '"'), ("&rdquo;", '"'), ("&mdash;", "—"), ("&ndash;", "–")):
        text = text.replace(a, b)
    text = _re.sub(r"[^\S\n]+", " ", text)
    text = _re.sub(r"\n\s*\n\s*\n+", "\n\n", text)
    return text.strip()


@app.get("/api/read-doc")
def read_doc(url: str) -> dict:
    """Fetch `url` and return its readable text so it can be spoken aloud.
    Returns {ok, kind, title, text, truncated} or {ok: False, error}."""
    import io as _io
    import re as _re
    from urllib.parse import urlparse

    try:
        parsed = urlparse(url)
    except Exception:
        return {"ok": False, "error": "bad-url"}
    if parsed.scheme not in ("http", "https") or not _read_doc_host_ok(parsed.hostname):
        return {"ok": False, "error": "blocked-url"}

    import requests  # already a dependency
    try:
        resp = requests.get(
            url, timeout=15, stream=True,
            headers={"User-Agent": "MinoBimaadiziwinAtlas/1.0 (+read-aloud)"},
        )
        resp.raise_for_status()
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": "fetch-failed", "detail": str(exc)[:200]}

    ctype = (resp.headers.get("content-type") or "").lower()
    # bounded read
    chunks, total = [], 0
    for chunk in resp.iter_content(64 * 1024):
        if not chunk:
            break
        chunks.append(chunk)
        total += len(chunk)
        if total > _READ_DOC_MAX_BYTES:
            break
    raw = b"".join(chunks)

    is_pdf = "pdf" in ctype or parsed.path.lower().endswith(".pdf") or raw[:5] == b"%PDF-"
    title = (parsed.path.rsplit("/", 1)[-1] or parsed.hostname or "document")

    if is_pdf:
        try:
            from pypdf import PdfReader
        except Exception:
            return {"ok": False, "error": "pdf-unsupported"}
        try:
            reader = PdfReader(_io.BytesIO(raw))
            if getattr(reader, "is_encrypted", False):
                try:
                    reader.decrypt("")
                except Exception:
                    pass
            parts = []
            for i, page in enumerate(reader.pages):
                if i >= _READ_DOC_MAX_PAGES:
                    break
                try:
                    parts.append(page.extract_text() or "")
                except Exception:
                    continue
            text = "\n".join(p.strip() for p in parts if p and p.strip())
            if reader.metadata and reader.metadata.title:
                title = str(reader.metadata.title)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": "pdf-parse-failed", "detail": str(exc)[:200]}
        kind = "pdf"
    else:
        try:
            html = raw.decode(resp.encoding or "utf-8", errors="replace")
        except Exception:
            html = raw.decode("utf-8", errors="replace")
        m = _re.search(r"(?is)<title[^>]*>(.*?)</title>", html)
        if m:
            title = _re.sub(r"\s+", " ", m.group(1)).strip() or title
        text = _html_to_text(html)
        kind = "web"

    text = " ".join((text or "").split())
    if not text:
        return {"ok": False, "error": "no-text",
                "message": "This document has no machine-readable text (it may be scanned images)."}
    truncated = len(text) > _READ_DOC_MAX_CHARS
    if truncated:
        text = text[:_READ_DOC_MAX_CHARS].rsplit(" ", 1)[0]
    return {"ok": True, "kind": kind, "title": title, "text": text, "truncated": truncated}


# ------------------------------ auth -------------------------------------- #

@app.post("/api/auth/login", response_model=LoginOut)
def login(response: Response,
          username: str = Form(...),
          password: str = Form(...)) -> LoginOut:
    user = db.verify_user(username, password)
    if not user:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Bad credentials")
    token = auth.create_access_token(user["username"], user["role"])
    response.set_cookie(
        "atlas_token", token,
        httponly=True, samesite="lax",
        max_age=config.JWT_EXPIRE_HOURS * 3600,
    )
    db.log_audit(user["username"], "login")
    return LoginOut(access_token=token, user=user)


@app.post("/api/auth/logout")
def logout(response: Response, user: dict = Depends(auth.optional_user)) -> dict:
    response.delete_cookie("atlas_token")
    if user:
        db.log_audit(user.get("username"), "logout")
    return {"ok": True}


@app.get("/api/auth/me")
def me(user: dict = Depends(auth.get_current_user)) -> dict:
    return user


@app.post("/api/auth/request-reset")
def auth_request_reset(payload: ResetRequestIn, request: Request) -> dict:
    """Generate a single-use password reset token.

    Always returns a 200 (no information leak). Behaviour:
    - if SMTP is configured AND the user has an email on file → email the link
    - otherwise → return `resetUrl` in the response so an admin can share it.
    """
    user = db.get_user(payload.username)
    if not user:
        # Don't reveal whether the user exists.
        return {"ok": True, "delivery": "n/a"}

    token = db.create_reset_token(user["username"])
    base = (config.PUBLIC_BASE_URL or str(request.base_url).rstrip("/"))
    reset_url = f"{base}/cms/reset.html?token={token}"

    delivery = "manual"
    if user.get("email") and mailer.is_configured():
        ok = mailer.send(
            user["email"],
            "Mino Bimaadiziwin CMS — password reset",
            f"Hello {user['username']},\n\nUse this link to reset your password "
            f"(expires in 1 hour):\n\n{reset_url}\n\nIf you didn't request this, ignore this email.",
        )
        delivery = "email" if ok else "manual"

    db.log_audit("system", "password.reset.request", user["username"], delivery)
    out: dict = {"ok": True, "delivery": delivery}
    # Only echo the URL back when we couldn't email it (admin needs to share it).
    if delivery != "email":
        out["resetUrl"] = reset_url
    return out


@app.post("/api/auth/reset")
def auth_reset(payload: ResetConfirmIn) -> dict:
    username = db.consume_reset_token(payload.token)
    if not username:
        raise HTTPException(400, "Invalid or expired token")
    ok = db.change_password(username, payload.password)
    if not ok:
        raise HTTPException(404, "User no longer exists")
    db.log_audit(username, "password.reset.complete")
    return {"ok": True}


# ------------------------------ users ------------------------------------- #

@app.get("/api/users")
def users_list(_: dict = Depends(auth.require_admin)) -> dict:
    return {"users": db.list_users()}


@app.post("/api/users", status_code=201)
def users_create(payload: UserIn, actor: dict = Depends(auth.require_admin)) -> dict:
    try:
        u = db.create_user(payload.username, payload.password, payload.role, payload.email)
    except Exception as e:
        raise HTTPException(400, f"Could not create user: {e}")
    db.log_audit(actor["username"], "user.create", payload.username)
    return u


@app.post("/api/users/{username}/email")
def users_set_email(username: str, payload: EmailIn,
                    actor: dict = Depends(auth.require_admin)) -> dict:
    ok = db.set_user_email(username, payload.email)
    if not ok:
        raise HTTPException(404, "User not found")
    db.log_audit(actor["username"], "user.email", username, payload.email or "(cleared)")
    return {"ok": True}


@app.delete("/api/users/{username}")
def users_delete(username: str, actor: dict = Depends(auth.require_admin)) -> dict:
    if username == actor["username"]:
        raise HTTPException(400, "Cannot delete yourself")
    ok = db.delete_user(username)
    if not ok:
        raise HTTPException(404, "User not found")
    db.log_audit(actor["username"], "user.delete", username)
    return {"ok": True}


@app.post("/api/users/{username}/password")
def users_change_password(username: str, payload: PasswordIn,
                          actor: dict = Depends(auth.require_admin)) -> dict:
    ok = db.change_password(username, payload.password)
    if not ok:
        raise HTTPException(404, "User not found")
    db.log_audit(actor["username"], "user.password", username)
    return {"ok": True}


# ---------------------------- communities --------------------------------- #

def _build_records() -> list[dict]:
    ds = db.current_dataset()
    base = ds["records"] if ds else []
    edits = db.latest_edits()
    enriched = processor.enrich(base)
    if not edits:
        return enriched
    out = []
    for r in enriched:
        cid = r.get("id")
        if cid and cid in edits:
            patch = edits[cid]
            merged = {**r, **{k: v for k, v in patch.items()
                              if k not in ("staff", "departments")}}
            if "staff" in patch:
                merged["staff"] = patch["staff"]
            if "departments" in patch:
                merged["departments"] = patch["departments"]
            # Recompute hasX flags after edits
            for key in analytics.PILLAR_KEYS + ["youth", "survivors", "connect"]:
                cap = f"has{key.capitalize()}"
                val = str(merged.get(key, "") or "").strip().lower()
                merged[cap] = val not in {"", "missing information", "needs review",
                                          "n/a", "no definite value", "duplicate record"}
            out.append(merged)
        else:
            out.append(r)
    return out


@app.get("/api/communities")
def communities_all() -> dict:
    """The dataset. If the database is unreachable we still answer, using the
    dataset bundled with this deploy, so the public dashboard keeps working."""
    try:
        records = coords.fill_missing_coords(_build_records())
        ds = db.current_dataset()
        return {
            "records": records,
            "count": len(records),
            "datasetVersion": ds["version"] if ds else 0,
            "datasetUploadedAt": ds["uploaded_at"] if ds else None,
            "datasetSource": ds["source_filename"] if ds else None,
        }
    except Exception as exc:                       # noqa: BLE001
        DB_OFFLINE["offline"] = True
        DB_OFFLINE["error"] = f"{type(exc).__name__}: {exc}"
        try:
            records = coords.fill_missing_coords(processor.load_initial_records() or [])
        except Exception:
            records = []
        return {
            "records": records,
            "count": len(records),
            "datasetVersion": 0,
            "datasetUploadedAt": None,
            "datasetSource": "communities-data.js (fallback — database offline)",
            "degraded": True,
        }


# ----- background AI enrichment job (auto-fills blanks from the web) -------- #
import threading as _threading
from . import llm as _llm

ENRICH_JOB = {"status": "idle", "done": 0, "total": 0, "started": None,
              "finished": None, "result": None, "error": None}


def _run_enrichment(use_llm: bool):
    ds = db.current_dataset()
    if not ds:
        ENRICH_JOB.update(status="error", error="No dataset uploaded yet.")
        return
    records = ds["records"]
    ENRICH_JOB.update(status="running", done=0, total=0, started=time.time(),
                      finished=None, result=None, error=None)

    def progress(done, total):
        ENRICH_JOB["done"], ENRICH_JOB["total"] = done, total

    try:
        summary = enrich.apply_enrichment(records, use_llm=use_llm, progress=progress)
        coords.fill_missing_coords(records)   # backfill any coords still missing
        db.replace_current_records(records)
        events.publish("dataset.enriched", summary)
        ENRICH_JOB.update(status="done", result=summary, finished=time.time())
    except Exception as e:  # pragma: no cover
        ENRICH_JOB.update(status="error", error=str(e), finished=time.time())


def _start_enrichment(use_llm: bool) -> bool:
    if ENRICH_JOB["status"] == "running":
        return False
    _threading.Thread(target=_run_enrichment, args=(use_llm,), daemon=True).start()
    return True


@app.post("/api/enrich/run")
def enrich_run(actor: dict = Depends(auth.require_admin)) -> dict:
    """Manually run the AI enrichment over the whole current dataset."""
    started = _start_enrichment(use_llm=_llm.available())
    if not started:
        raise HTTPException(409, "Enrichment is already running.")
    return {"ok": True, "using_llm": _llm.available()}


@app.get("/api/enrich/status")
def enrich_status(actor: dict = Depends(auth.require_admin)) -> dict:
    return {**ENRICH_JOB, "llm_available": _llm.available()}


class GroqKeyIn(BaseModel):
    key: str = ""


@app.post("/api/enrich/groq-key")
def enrich_groq_key(body: GroqKeyIn, actor: dict = Depends(auth.require_admin)) -> dict:
    """Save the team's free Groq API key (stored in settings, used to structure
    scraped contact info). Pass an empty string to clear it."""
    db.set_setting("ai.groq_key", (body.key or "").strip(), actor["username"])
    _llm.set_key((body.key or "").strip())
    db.log_audit(actor["username"], "ai.groq_key", None, "set" if body.key else "cleared")
    return {"ok": True, "llm_available": _llm.available()}


@app.get("/api/enrich/community")
def enrich_community(name: str, actor: dict = Depends(auth.require_admin)) -> dict:
    """Run the free AI/data check for ONE community (fast). Returns proposed
    corrections — population, official website, coordinates, scraped contacts
    and document links — each with a source URL and confidence. Never writes
    anything; the CMS shows them for a human to accept."""
    records = _build_records()
    rec = next((r for r in records if (r.get("name") or "").strip().lower() == name.strip().lower()), None)
    cur_link = (rec or {}).get("link") or (rec or {}).get("website") or ""
    cur_pop = (rec or {}).get("population")
    try:
        suggestions = enrich.enrich_one(name, cur_link, cur_pop, scrape=True, use_llm=_llm.available())
    except Exception as e:
        raise HTTPException(502, f"Lookup failed: {e}")
    return {"community": name, "suggestions": suggestions}


@app.get("/api/export/xlsx")
def export_xlsx(actor: dict = Depends(auth.require_admin)):
    """Download the current (CMS-corrected) dataset as an Excel workbook."""
    import io
    import openpyxl
    from openpyxl.styles import Font as _Font, PatternFill as _Fill
    records = coords.fill_missing_coords(_build_records())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Communities"
    cols = [("Community", "name"), ("Direction", "direction"), ("Type", "orgType"),
            ("Population", "population"), ("Latitude", "lat"), ("Longitude", "lng"),
            ("Link", "link"), ("Contact", "contact"),
            ("Physical", "physical"), ("Mental", "mental"),
            ("Spiritual", "spiritual"), ("Emotional", "emotional"),
            ("Survivors", "survivors"), ("Youth", "youth"),
            ("Strategic Plan", "strategicPlan"), ("AGM", "agm"), ("Financials", "financials")]
    ws.append([c[0] for c in cols])
    for i in range(1, len(cols) + 1):
        ws.cell(1, i).font = _Font(bold=True, color="FFFFFF")
        ws.cell(1, i).fill = _Fill("solid", fgColor="1F2A37")
    from openpyxl.comments import Comment as _Comment
    ai_fill = _Fill("solid", fgColor="C7B3F0")   # purple = AI-pulled
    link_font = _Font(color="5B3A9E", underline="single")
    for r in records:
        lng = r.get("lng") if r.get("lng") is not None else r.get("lon")
        ws.append([r.get(k) if k != "lng" else lng for (_, k) in cols])
        ai = r.get("_ai") or {}
        notes = r.get("_ai_note") or {}
        if ai:
            row_i = ws.max_row
            for col_i, (_, k) in enumerate(cols, 1):
                if k in ai:
                    cell = ws.cell(row_i, col_i)
                    cell.fill = ai_fill
                    val = str(cell.value or "")
                    if val.startswith("http"):           # make AI links clickable
                        cell.hyperlink = val
                        cell.font = link_font
                    expl = notes.get(k)                   # AI explanation as a comment
                    if expl:
                        cell.comment = _Comment(f"AI: {expl}", "AI")
    ws.freeze_panes = "A2"
    # legend note
    ws2 = wb.create_sheet("Legend")
    ws2["A1"] = "Cell colour"; ws2["B1"] = "Meaning"
    ws2["A1"].font = ws2["B1"].font = _Font(bold=True)
    ws2["A2"].fill = ai_fill
    ws2["B2"] = "AI Suggested (verify) — auto-filled from a free web source"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = "mino-bimaadiziwin-atlas.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"',
                 "Cache-Control": "no-store"},
    )


@app.get("/api/communities/edits")
def communities_edits(_: dict = Depends(auth.require_admin)) -> dict:
    return {"edits": db.latest_edits(), "history": db.edit_history(limit=100)}


@app.get("/api/communities/versions")
def communities_versions(_: dict = Depends(auth.require_admin)) -> dict:
    return {"versions": db.list_dataset_versions()}


@app.post("/api/communities/versions/{version}/activate")
def communities_activate(version: int, actor: dict = Depends(auth.require_admin)) -> dict:
    ok = db.activate_dataset_version(version)
    if not ok:
        raise HTTPException(404, "Version not found")
    db.log_audit(actor["username"], "dataset.activate", str(version))
    # Refresh on-disk JS file too so static dashboard sees it
    ds = db.current_dataset()
    if ds:
        processor.write_communities_js(ds["records"])
    events.publish("dataset.updated", {"version": version, "activated": True})
    return {"ok": True, "version": version}


@app.post("/api/communities/upload")
async def communities_upload(
    file: UploadFile = File(...),
    actor: dict = Depends(auth.require_admin),
) -> dict:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in (".xlsx", ".xls", ".csv"):
        raise HTTPException(400, "Only .xlsx, .xls, or .csv files are accepted")
    target = config.UPLOADS_DIR / f"{uuid.uuid4().hex}{suffix}"
    with target.open("wb") as f:
        shutil.copyfileobj(file.file, f)

    previous = (db.current_dataset() or {}).get("records") or []
    snapshot: dict | None = None
    try:
        # State-of-the-art path: full multi-sheet ingestion (Master Sheet +
        # 85-list + Missing/Review/Correction + Additional Links). Only
        # works for .xlsx — for .csv we fall back to the simple processor.
        if suffix in (".xlsx", ".xls"):
            snapshot = workbook.ingest_workbook(target)
            records = snapshot["records"]
            # Carry forward lat/lon from previous dataset where missing
            by_name = {processor._slugify(r.get("name", "")): r for r in previous}
            for r in records:
                if r.get("lat") is None or r.get("lon") is None:
                    prior = by_name.get(processor._slugify(r.get("name", "")))
                    if prior:
                        r.setdefault("lat", prior.get("lat"))
                        r.setdefault("lon", prior.get("lon"))
        else:
            records = processor.process_file(target, previous=previous)
    except Exception as e:
        raise HTTPException(400, f"Processing failed: {e}")

    if not records:
        raise HTTPException(400, "No records parsed from file")

    # fill in coordinates the sheet omits (real communities + org HQs)
    coords.fill_missing_coords(records)

    version = db.save_dataset(
        records,
        source_filename=file.filename,
        uploaded_by=actor["username"],
        snapshot=snapshot,
    )
    processor.write_communities_js(records)
    db.log_audit(actor["username"], "dataset.upload",
                 file.filename, f"v{version}, {len(records)} records")
    events.publish("dataset.updated", {
        "version": version,
        "records": len(records),
        "source": file.filename,
        "annotations": (snapshot or {}).get("annotations", {}).get("totals", {}),
    })
    # Kick off AI enrichment automatically in the background — fills blank
    # objective cells from the web and colour-codes them as AI. The upload
    # response returns immediately; the CMS polls /api/enrich/status.
    _start_enrichment(use_llm=_llm.available())

    out = {"ok": True, "version": version, "records": len(records), "enriching": True}
    if snapshot:
        out["meta"] = snapshot["meta"]
        out["annotations"] = snapshot["annotations"]["totals"]
    return out


class GSheetSyncIn(BaseModel):
    spreadsheetId: str | None = None
    apiKey: str | None = None


@app.post("/api/communities/sync-gsheet")
def communities_sync_gsheet(payload: GSheetSyncIn,
                            actor: dict = Depends(auth.require_admin)) -> dict:
    """Pull the dataset straight from the live Google Sheet via the Sheets API.
    Unlike .xlsx upload, this preserves EVERY in-cell link. Credentials resolve
    from the request body, then env (GOOGLE_SHEETS_ID / GOOGLE_SHEETS_API_KEY),
    then CMS settings (sheets.spreadsheetId / sheets.apiKey)."""
    settings = db.get_settings()
    sid = (payload.spreadsheetId or os.environ.get("GOOGLE_SHEETS_ID", "")
           or settings.get("sheets.spreadsheetId", "")).strip()
    key = (payload.apiKey or os.environ.get("GOOGLE_SHEETS_API_KEY", "")
           or settings.get("sheets.apiKey", "")).strip()
    if not sid or not key:
        raise HTTPException(400, "Provide a Google Sheet ID/URL and an API key "
                                 "(or set them in Settings / env vars).")
    # Persist for next time (so the admin can just click 'Sync' later).
    try:
        db.set_setting("sheets.spreadsheetId", gsheets.extract_spreadsheet_id(sid), actor["username"])
        if not os.environ.get("GOOGLE_SHEETS_API_KEY"):
            db.set_setting("sheets.apiKey", key, actor["username"])
    except Exception:
        pass

    previous = (db.current_dataset() or {}).get("records") or []
    try:
        snapshot = gsheets.ingest_gsheet(sid, key)
        records = snapshot["records"]
        by_name = {processor._slugify(r.get("name", "")): r for r in previous}
        for r in records:
            if r.get("lat") is None or r.get("lon") is None:
                prior = by_name.get(processor._slugify(r.get("name", "")))
                if prior:
                    r.setdefault("lat", prior.get("lat"))
                    r.setdefault("lon", prior.get("lon"))
    except Exception as e:
        raise HTTPException(400, f"Google Sheets sync failed: {e}")
    if not records:
        raise HTTPException(400, "No records parsed from the Google Sheet")

    coords.fill_missing_coords(records)
    version = db.save_dataset(records, source_filename="Google Sheet (live sync)",
                              uploaded_by=actor["username"], snapshot=snapshot)
    processor.write_communities_js(records)
    db.log_audit(actor["username"], "dataset.gsheet-sync",
                 "Google Sheet", f"v{version}, {len(records)} records")
    events.publish("dataset.updated", {
        "version": version, "records": len(records), "source": "Google Sheet (live sync)",
        "annotations": (snapshot or {}).get("annotations", {}).get("totals", {}),
    })
    link_total = sum(len(v) for r in records for v in (r.get("fieldLinks") or {}).values())
    return {"ok": True, "version": version, "records": len(records),
            "fieldLinks": link_total, "meta": snapshot.get("meta")}


@app.post("/api/communities/import-gscript")
async def communities_import_gscript(request: Request) -> dict:
    """Receive sheet data (with ALL in-cell links) from the Google Apps Script
    that runs inside the user's sheet. Auth is a shared token so it works
    without a login flow. Works even when the org blocks link-sharing and
    service accounts, because the script runs as the signed-in user."""
    expected = (os.environ.get("GSCRIPT_SYNC_TOKEN", "")
                or db.get_settings().get("sync.gscriptToken", "")).strip()
    if not expected:
        raise HTTPException(400, "Apps Script sync is not configured yet "
                                 "(set GSCRIPT_SYNC_TOKEN).")
    token = (request.headers.get("X-Sync-Token") or "").strip()
    if token != expected:
        raise HTTPException(401, "Bad sync token")
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON body")

    previous = (db.current_dataset() or {}).get("records") or []
    try:
        snapshot = gsheets.ingest_gscript(payload)
        records = snapshot["records"]
        by_name = {processor._slugify(r.get("name", "")): r for r in previous}
        for r in records:
            if r.get("lat") is None or r.get("lon") is None:
                prior = by_name.get(processor._slugify(r.get("name", "")))
                if prior:
                    r.setdefault("lat", prior.get("lat"))
                    r.setdefault("lon", prior.get("lon"))
    except Exception as e:
        raise HTTPException(400, f"Apps Script import failed: {e}")
    if not records:
        raise HTTPException(400, "No records parsed from the Apps Script payload")

    coords.fill_missing_coords(records)
    version = db.save_dataset(records, source_filename="Google Sheet (Apps Script sync)",
                              uploaded_by="apps-script", snapshot=snapshot)
    processor.write_communities_js(records)
    db.log_audit("apps-script", "dataset.gscript-sync", "Google Sheet",
                 f"v{version}, {len(records)} records")
    events.publish("dataset.updated", {
        "version": version, "records": len(records),
        "source": "Google Sheet (Apps Script sync)",
        "annotations": (snapshot or {}).get("annotations", {}).get("totals", {}),
    })
    link_total = sum(len(v) for r in records for v in (r.get("fieldLinks") or {}).values())
    return {"ok": True, "version": version, "records": len(records), "fieldLinks": link_total}


@app.post("/api/communities/{community_id}/edit")
def communities_edit(community_id: str, payload: EditIn,
                     actor: dict = Depends(auth.require_admin)) -> dict:
    merged = dict(payload.fields)
    if payload.staff is not None:
        merged["staff"] = payload.staff
    if payload.departments is not None:
        merged["departments"] = payload.departments
    edit_id = db.save_edit(community_id, merged, actor["username"])
    db.log_audit(actor["username"], "community.edit", community_id,
                 f"fields: {', '.join(merged.keys())}")
    events.publish("community.edited", {"id": community_id, "fields": list(merged.keys())})
    return {"ok": True, "editId": edit_id}


@app.delete("/api/communities/{community_id}")
def communities_delete(community_id: str, actor: dict = Depends(auth.require_admin)) -> dict:
    db.save_edit(community_id, {"_deleted": True}, actor["username"])
    db.log_audit(actor["username"], "community.delete", community_id)
    events.publish("community.deleted", {"id": community_id})
    return {"ok": True}


# ----------------------- workbook cross-references ------------------------ #

@app.get("/api/communities/coverage85")
def coverage85() -> dict:
    """85-community canonical list × whether each appears in the master sheet."""
    snap = db.current_snapshot()
    if not snap:
        return {"items": [], "summary": {"total": 0, "covered": 0, "missing": 0}}
    items = snap.get("coverage85", [])
    covered = sum(1 for c in items if c.get("inMaster"))
    return {
        "items": items,
        "extras": snap.get("list85", {}).get("inMasterNotInList", []),
        "summary": {
            "total": len(items),
            "covered": covered,
            "missing": len(items) - covered,
        },
    }


@app.get("/api/communities/annotations")
def workbook_annotations(community: str | None = None) -> dict:
    """Missing-info / Needs-review / Correction items, optionally filtered."""
    snap = db.current_snapshot()
    if not snap:
        return {"missing": [], "review": [], "corrections": [], "totals": {}}
    ann = snap.get("annotations", {})
    if community:
        key = workbook.normalize(community)
        return {
            "missing": [a for a in ann.get("missing", []) if a.get("communityKey") == key],
            "review": [a for a in ann.get("review", []) if a.get("communityKey") == key],
            "corrections": [a for a in ann.get("corrections", []) if a.get("communityKey") == key],
        }
    return {
        "missing": ann.get("missing", []),
        "review": ann.get("review", []),
        "corrections": ann.get("corrections", []),
        "totals": ann.get("totals", {}),
        "additionalLinks": snap.get("additionalLinks", []),
        "meta": snap.get("meta", {}),
    }


# ----------------------------- analytics ---------------------------------- #

@app.get("/api/analytics/overview")
def an_overview() -> dict:
    return analytics.overview(_build_records())


@app.get("/api/analytics/pillars")
def an_pillars() -> dict:
    return analytics.pillars_breakdown(_build_records())


@app.get("/api/analytics/gaps")
def an_gaps() -> dict:
    return analytics.gaps(_build_records())


@app.get("/api/analytics/population")
def an_population() -> dict:
    return analytics.population_distribution(_build_records())


@app.get("/api/analytics/keywords")
def an_keywords() -> dict:
    return analytics.keywords(_build_records())


@app.get("/api/analytics/clusters")
def an_clusters(k: int = 5) -> dict:
    return analytics.cluster_communities(_build_records(), n_clusters=k)


@app.get("/api/analytics/coverage")
def an_coverage() -> dict:
    return analytics.coverage_matrix(_build_records())


@app.get("/api/analytics/quality")
def an_quality() -> dict:
    return analytics.quality_report(_build_records())


@app.get("/api/analytics/full")
def an_full() -> dict:
    return analytics.full_report(_build_records())


@app.get("/api/analytics/duplicates")
def an_duplicates(min_similarity: float = 0.45) -> dict:
    return analytics.detect_duplicate_services(_build_records(), min_similarity=min_similarity)


@app.get("/api/analytics/search")
def an_search(q: str, limit: int = 8) -> dict:
    """Free retrieval-based 'chatbot' — TF-IDF over the master sheet,
    cites communities, no external API calls, no usage costs."""
    return analytics.semantic_search(_build_records(), q, limit=limit)


class AskIn(BaseModel):
    question: str = Field(min_length=1, max_length=600)


@app.post("/api/analytics/ask")
def an_ask(body: AskIn) -> dict:
    """AI research assistant, GROUNDED in the sheet. We compute the real
    numbers server-side (overview + data-quality report) and retrieve the most
    relevant sheet passages, then ask the configured LLM (Groq or xAI Grok —
    same settings box, key auto-detected) to answer ONLY from that evidence,
    citing communities and flagging data gaps. With no key configured we fall
    back to the pure-retrieval answer, so the box always works."""
    import json
    from . import llm
    records = _build_records()
    q = body.question.strip()
    retrieval = analytics.semantic_search(records, q, limit=6)
    if not llm.available():
        retrieval["ai"] = False
        return retrieval

    ov = analytics.overview(records)
    quality = analytics.quality_report(records)
    digest = {
        "communities_total": ov.get("total"),
        "by_direction": ov.get("byDirection"),
        "pillars_documented_counts": ov.get("pillarsCovered"),
        "all_four_pillars": ov.get("hasAllPillars"),
        "population_documented_for": ov.get("populationKnown"),
        "population_total_where_documented": ov.get("populationTotal"),
        "avg_field_completeness_0to1": ov.get("completenessScore"),
        "data_quality_issues": quality,
    }
    snippets = "\n".join(
        f"- {r.get('name')} ({r.get('direction') or '—'}): {r.get('snippet')}"
        for r in (retrieval.get("results") or [])[:6]
    )
    # If the question NAMES a community, hand the model that community's FULL
    # row (every field, trimmed) — so asking about one place gets a real,
    # drawer-level answer instead of a one-line snippet.
    ql = q.lower()
    full_rows = []
    for rec in records:
        nm = str(rec.get("name") or "").strip()
        if len(nm) >= 5 and nm.lower() in ql:
            fields = {}
            for fk in ("direction", "type", "population", "physical", "mental", "spiritual",
                       "emotional", "youth", "survivors", "connect", "contact",
                       "strategicPlan", "agm", "financials"):
                fv = str(rec.get(fk) or "").strip()
                if fv:
                    fields[fk] = fv[:900]
            full_rows.append({"name": nm, **fields})
        if len(full_rows) >= 2:
            break
    full_block = ""
    if full_rows:
        full_block = ("\nFULL SHEET ROW(S) for the community/communities named in the question "
                      "(answer from these in detail):\n" + json.dumps(full_rows, default=str) + "\n")
    system = (
        "You are the research assistant for the Mino Bimaadiziwin Community "
        "Services Atlas — First Nations community health & wellness data. "
        "Answer ONLY from the DATA DIGEST and SHEET PASSAGES provided. Never "
        "invent numbers, programs, or communities. Cite community names when "
        "you use their passages. The sheet has known gaps — when the digest's "
        "data_quality_issues affect the answer (missing population, "
        "needs-review fields), say so plainly in one short sentence. Write "
        "warmly and simply (elders read this). 120 words max. Plain text only — "
        "no markdown headers or bullets unless listing communities."
    )
    user = (
        f"DATA DIGEST (computed live from the master sheet):\n{json.dumps(digest, default=str)}\n"
        f"{full_block}\n"
        f"SHEET PASSAGES most relevant to the question:\n{snippets or '(none found)'}\n\n"
        f"QUESTION: {q}"
    )
    answer = llm.chat(system, user, max_tokens=520, temperature=0.2)
    if not answer:
        retrieval["ai"] = False
        return retrieval
    return {"answer": answer, "results": (retrieval.get("results") or [])[:4],
            "ai": True, "provider": llm.provider_label()}


@app.get("/api/analytics/compare")
def an_compare(regions: str | None = None) -> dict:
    region_list = [r.strip() for r in regions.split(",")] if regions else None
    return analytics.compare_regions(_build_records(), region_list)


@app.get("/api/analytics/facts")
def an_facts() -> dict:
    return analytics.storytelling_facts(_build_records())


@app.get("/api/analytics/peers")
def an_peers(name: str, k: int = 6) -> dict:
    return analytics.find_peers(_build_records(), name, k=k)


@app.get("/api/analytics/gaps_for")
def an_gaps_for(name: str) -> dict:
    return analytics.find_gaps_for(_build_records(), name)


@app.get("/api/analytics/spotlight")
def an_spotlight(name: str) -> dict:
    return analytics.spotlight(_build_records(), name)


@app.get("/api/audit")
def audit_log(_: dict = Depends(auth.require_admin), limit: int = 100) -> dict:
    return {"events": db.recent_audit(limit)}


# ----------------------------- settings ----------------------------------- #

@app.get("/api/settings")
def settings_get() -> dict:
    """Public — the dashboard reads these to customise the UI. Secrets
    (API keys / tokens) are NEVER returned here; they are write-only from the
    admin UI and live only server-side (DB or env)."""
    return _redact_secrets(db.get_settings())


# Setting keys that must never be exposed on the public settings endpoint.
_SECRET_SETTING_KEYS = {"sheets.apiKey", "ai.groq_key"}


def _redact_secrets(settings: dict) -> dict:
    out = {}
    for k, v in (settings or {}).items():
        lk = str(k).lower()
        if k in _SECRET_SETTING_KEYS or lk.endswith(("apikey", "secret", "token", "password")) or lk.endswith(".key"):
            continue  # drop entirely — never leak
        out[k] = v
    return out


@app.put("/api/settings")
def settings_update(payload: SettingsIn, actor: dict = Depends(auth.require_admin)) -> dict:
    db.bulk_set_settings(payload.values, actor["username"])
    db.log_audit(actor["username"], "settings.update", None,
                 f"keys: {', '.join(payload.values.keys())}")
    events.publish("settings.updated", {"keys": list(payload.values.keys())})
    return _redact_secrets(db.get_settings())


@app.post("/api/settings/reset")
def settings_reset(actor: dict = Depends(auth.require_admin)) -> dict:
    """Reset all site.* settings back to their first-boot defaults."""
    db.bulk_set_settings(db.DEFAULT_SETTINGS, actor["username"])
    db.log_audit(actor["username"], "settings.reset", None,
                 f"reset {len(db.DEFAULT_SETTINGS)} keys to defaults")
    events.publish("settings.updated", {"keys": list(db.DEFAULT_SETTINGS.keys())})
    return db.get_settings()


# ----------------------------- pages -------------------------------------- #

@app.get("/api/pages")
def pages_list(visible_only: bool = False) -> dict:
    return {"pages": db.list_pages(only_visible=visible_only)}


@app.get("/api/pages/{slug}")
def pages_get(slug: str) -> dict:
    p = db.get_page(slug)
    if not p:
        raise HTTPException(404, "Page not found")
    return p


@app.put("/api/pages/{slug}")
def pages_upsert(slug: str, payload: PageIn,
                 actor: dict = Depends(auth.require_admin)) -> dict:
    p = db.upsert_page(payload.slug or slug, payload.title, payload.body,
                       payload.visible, actor["username"])
    db.log_audit(actor["username"], "page.upsert", payload.slug or slug)
    events.publish("page.updated", {"slug": payload.slug or slug})
    return p


@app.delete("/api/pages/{slug}")
def pages_delete(slug: str, actor: dict = Depends(auth.require_admin)) -> dict:
    ok = db.delete_page(slug)
    if not ok:
        raise HTTPException(404, "Page not found")
    db.log_audit(actor["username"], "page.delete", slug)
    events.publish("page.deleted", {"slug": slug})
    return {"ok": True}


# --------------------------- navigation items ---------------------------- #
# Lets the admin reorder/rename/hide/show the dashboard's top navigation.

@app.get("/api/nav")
def nav_list(slot: str | None = None) -> dict:
    return {"items": db.list_nav(slot=slot)}


@app.put("/api/nav")
def nav_upsert(payload: NavItemIn, actor: dict = Depends(auth.require_admin)) -> dict:
    item = db.upsert_nav(payload.model_dump(), actor["username"])
    db.log_audit(actor["username"], "nav.upsert", payload.view)
    events.publish("nav.updated", {"view": payload.view})
    return item


@app.delete("/api/nav/{nav_id}")
def nav_delete(nav_id: int, actor: dict = Depends(auth.require_admin)) -> dict:
    ok = db.delete_nav(nav_id)
    if not ok:
        raise HTTPException(404, "Nav item not found")
    db.log_audit(actor["username"], "nav.delete", str(nav_id))
    events.publish("nav.updated", {"id": nav_id, "deleted": True})
    return {"ok": True}


# ----------------------------- layouts ------------------------------------ #
# Layouts are ordered lists of "blocks" that the dashboard renders above
# (slot=header) and below (slot=footer) the view-switched main content.
# Admin can drag-and-drop reorder, toggle visibility, and add custom blocks.

@app.get("/api/layouts")
def layouts_all() -> dict:
    """Public — the dashboard reads these on every page load."""
    return db.get_all_layouts()


@app.get("/api/layouts/{name}")
def layouts_get(name: str) -> dict:
    return {"name": name, "blocks": db.get_layout(name)}


@app.put("/api/layouts/{name}")
def layouts_set(name: str, payload: LayoutIn,
                actor: dict = Depends(auth.require_admin)) -> dict:
    blocks = db.save_layout(name, payload.blocks, actor["username"])
    db.log_audit(actor["username"], "layout.update", name,
                 f"{len(blocks)} blocks")
    events.publish("layout.updated", {"name": name})
    return {"name": name, "blocks": blocks}


@app.post("/api/layouts/{name}/reset")
def layouts_reset(name: str, actor: dict = Depends(auth.require_admin)) -> dict:
    blocks = db.reset_layout(name, actor["username"])
    db.log_audit(actor["username"], "layout.reset", name)
    events.publish("layout.updated", {"name": name, "reset": True})
    return {"name": name, "blocks": blocks}


# --------------------------- element styles ------------------------------- #

@app.get("/api/styles")
def styles_all() -> dict:
    """Public — dashboard reads these on every page load to inject CSS."""
    return db.list_styles()


@app.get("/api/styles/{selector:path}")
def styles_get(selector: str) -> dict:
    s = db.get_style(selector)
    if not s:
        return {"selector": selector, "properties": {}, "customCss": ""}
    return s


@app.put("/api/styles/{selector:path}")
def styles_set(selector: str, payload: StyleIn,
               actor: dict = Depends(auth.require_admin)) -> dict:
    s = db.save_style(selector, payload.properties, payload.customCss, actor["username"])
    db.log_audit(actor["username"], "style.update", selector,
                 f"{len(payload.properties)} props")
    events.publish("style.updated", {"selector": selector})
    return s


@app.delete("/api/styles/{selector:path}")
def styles_delete(selector: str, actor: dict = Depends(auth.require_admin)) -> dict:
    ok = db.delete_style(selector)
    if not ok:
        raise HTTPException(404, "Style not found")
    db.log_audit(actor["username"], "style.delete", selector)
    events.publish("style.updated", {"selector": selector, "deleted": True})
    return {"ok": True}


@app.post("/api/styles/_reset_all")
def styles_reset_all(actor: dict = Depends(auth.require_admin)) -> dict:
    """Nuke every element_styles row — the panic button for when admins
    have made things invisible through accumulated overrides."""
    with db.get_conn() as conn:
        from sqlalchemy import text as _t
        n = conn.execute(_t("SELECT COUNT(*) FROM element_styles")).scalar() or 0
        conn.execute(_t("DELETE FROM element_styles"))
    db.log_audit(actor["username"], "style.reset_all", None, f"{n} overrides cleared")
    events.publish("style.updated", {"resetAll": True})
    return {"ok": True, "cleared": int(n)}


@app.post("/api/styles/_reset_sizes")
def styles_reset_sizes(actor: dict = Depends(auth.require_admin)) -> dict:
    """Selective reset — strips ONLY dimensional properties (fontSize,
    width, height, transform) from every element_styles row. Preserves
    colours, padding/margin tweaks, and other styling. Use when admins
    have accidentally drag-resized things to wrong sizes but don't want
    to lose all their other style work."""
    import json as _j
    SIZE_KEYS = {"fontSize", "width", "height", "minWidth", "maxWidth",
                 "minHeight", "maxHeight", "transform"}

    def strip(obj):
        """Recursively delete size-related properties. Returns count removed."""
        n = 0
        if isinstance(obj, dict):
            for k in list(obj.keys()):
                v = obj[k]
                if k in SIZE_KEYS:
                    del obj[k]; n += 1; continue
                if isinstance(v, (dict, list)):
                    n += strip(v)
        elif isinstance(obj, list):
            for x in obj:
                n += strip(x)
        return n

    cleared = 0
    rows_updated = 0
    rows_deleted = 0
    with db.get_conn() as conn:
        from sqlalchemy import text as _t
        rows = conn.execute(_t("SELECT selector, properties FROM element_styles")).fetchall()
        for sel, props_json in rows:
            try:
                props = _j.loads(props_json or "{}")
            except Exception:
                continue
            removed = strip(props)
            if removed == 0:
                continue
            cleared += removed
            # If the row has no properties left, delete it; otherwise update.
            def empty(o):
                if isinstance(o, dict):
                    return all(empty(v) for v in o.values())
                if isinstance(o, list):
                    return all(empty(v) for v in o)
                return o in (None, "", {}, [])
            if empty(props):
                conn.execute(_t("DELETE FROM element_styles WHERE selector = :s"),
                             {"s": sel})
                rows_deleted += 1
            else:
                conn.execute(
                    _t("UPDATE element_styles SET properties = :p, "
                       "updated_at = :u, updated_by = :by WHERE selector = :s"),
                    {"p": _j.dumps(props), "u": __import__("time").time(),
                     "by": actor["username"], "s": sel},
                )
                rows_updated += 1
    db.log_audit(actor["username"], "style.reset_sizes", None,
                 f"{cleared} size props removed, "
                 f"{rows_updated} rows updated, {rows_deleted} rows deleted")
    events.publish("style.updated", {"resetSizes": True})
    return {"ok": True, "cleared": cleared,
            "rowsUpdated": rows_updated, "rowsDeleted": rows_deleted}


# ----------------------------- SSE channel -------------------------------- #

@app.get("/api/events")
async def event_stream() -> StreamingResponse:
    """Server-Sent Events — clients reconnect automatically.
    Emits dataset.updated, community.edited, settings.updated, page.updated, etc.
    """
    async def gen():
        async for chunk in events.subscribe():
            yield chunk
    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


# --------------------------- static hosting ------------------------------- #
# Serve the dashboard files at /, the CMS at /cms.

DASHBOARD_ROOT = config.ROOT
CMS_ROOT = config.ROOT / "cms"


@app.get("/")
def root_page() -> FileResponse:
    # no-cache forces the browser to revalidate the HTML on every load, so new
    # ?v= asset references (the cache-busted jsx/css) are always picked up.
    # Without this the browser keeps stale HTML and "nothing changes" on deploy.
    return FileResponse(
        DASHBOARD_ROOT / "Community Atlas.html",
        headers={"Cache-Control": "no-cache, must-revalidate"},
    )


@app.get("/favicon.ico")
def favicon() -> FileResponse:
    """Browsers still request /favicon.ico — serve the SVG so we don't 404."""
    return FileResponse(DASHBOARD_ROOT / "favicon.svg", media_type="image/svg+xml")


# ============================ MEDIA LIBRARY ============================== #

MEDIA_MAX_BYTES = 5 * 1024 * 1024  # 5 MB per file
MEDIA_ALLOWED_TYPES = {
    "image/jpeg", "image/png", "image/webp", "image/gif",
    "image/svg+xml", "image/avif",
}


@app.post("/api/media")
async def media_upload(
    file: UploadFile = File(...),
    alt: str = Form(""),
    actor: dict = Depends(auth.require_admin),
) -> dict:
    import base64
    content = await file.read()
    if len(content) > MEDIA_MAX_BYTES:
        raise HTTPException(400, f"File too large (max {MEDIA_MAX_BYTES // 1024} KB)")
    mime = file.content_type or "application/octet-stream"
    if mime not in MEDIA_ALLOWED_TYPES:
        raise HTTPException(400, f"Unsupported media type: {mime}")
    b64 = base64.b64encode(content).decode("ascii")
    row = db.save_media(file.filename or "upload", mime, b64, len(content),
                        alt or None, actor["username"])
    db.log_audit(actor["username"], "media.upload", file.filename,
                 f"{len(content)} bytes, {mime}")
    events.publish("media.updated", {"id": row.get("id")})
    return row


@app.get("/api/media")
def media_list(_: dict = Depends(auth.require_admin)) -> dict:
    return {"items": db.list_media()}


@app.get("/api/media/{media_id}")
def media_get(media_id: int):
    """Public — serve the raw image bytes so <img src="/api/media/12"> works."""
    import base64
    m = db.get_media_content(media_id)
    if not m:
        raise HTTPException(404, "Not found")
    try:
        raw = base64.b64decode(m["content_b64"])
    except Exception:
        raise HTTPException(500, "Corrupt media")
    return Response(content=raw, media_type=m.get("mime_type") or "application/octet-stream",
                    headers={"Cache-Control": "public, max-age=86400"})


@app.delete("/api/media/{media_id}")
def media_delete(media_id: int, actor: dict = Depends(auth.require_admin)) -> dict:
    ok = db.delete_media(media_id)
    if not ok:
        raise HTTPException(404, "Not found")
    db.log_audit(actor["username"], "media.delete", str(media_id))
    events.publish("media.updated", {"id": media_id, "deleted": True})
    return {"ok": True}


# ============================ SUBMISSIONS ================================ #

class SubmissionIn(BaseModel):
    form: str = Field(min_length=1, max_length=80)
    fields: dict


@app.post("/api/submissions")
def submission_create(request: Request, payload: SubmissionIn) -> dict:
    """Public — visitors POST here from form blocks on the dashboard."""
    # Trim to keep storage sane
    safe_fields = {
        str(k)[:80]: str(v)[:4000] for k, v in (payload.fields or {}).items()
    }
    ip = request.client.host if request.client else None
    ua = request.headers.get("user-agent", "")[:300]
    sub_id = db.save_submission(payload.form, safe_fields, ip, ua)
    events.publish("submission.received", {"form": payload.form, "id": sub_id})
    # Optional email notification (uses the same SMTP env vars as password reset)
    if mailer.is_configured():
        try:
            body_lines = [f"New submission to form '{payload.form}':", ""]
            for k, v in safe_fields.items():
                body_lines.append(f"  {k}: {v}")
            body_lines.append("")
            body_lines.append(f"From IP: {ip}, UA: {ua}")
            mailer.send(config.SMTP_FROM, f"New form submission: {payload.form}",
                        "\n".join(body_lines))
        except Exception:
            pass
    return {"ok": True, "id": sub_id}


@app.get("/api/submissions")
def submissions_list(form: str | None = None,
                     _: dict = Depends(auth.require_admin)) -> dict:
    return {"items": db.list_submissions(form_slug=form),
            "summary": db.submissions_summary()}


@app.post("/api/submissions/{sub_id}/read")
def submission_mark_read(sub_id: int,
                         _: dict = Depends(auth.require_admin)) -> dict:
    ok = db.mark_submission_read(sub_id)
    if not ok:
        raise HTTPException(404, "Not found")
    return {"ok": True}


@app.delete("/api/submissions/{sub_id}")
def submission_delete(sub_id: int,
                      actor: dict = Depends(auth.require_admin)) -> dict:
    ok = db.delete_submission(sub_id)
    if not ok:
        raise HTTPException(404, "Not found")
    db.log_audit(actor["username"], "submission.delete", str(sub_id))
    return {"ok": True}


# ============================ SEO ROUTES ================================= #

@app.get("/robots.txt")
def robots() -> Response:
    s = db.get_settings()
    allow = str(s.get("site.robotsAllow", "true")).lower() != "false"
    body = "User-agent: *\n"
    body += ("Allow: /\n" if allow else "Disallow: /\n")
    body += "Disallow: /api/\n"
    body += "Disallow: /cms\n"
    body += f"Sitemap: {(s.get('site.canonicalUrl') or config.PUBLIC_BASE_URL or '').rstrip('/')}/sitemap.xml\n"
    return Response(content=body, media_type="text/plain")


@app.get("/sitemap.xml")
def sitemap() -> Response:
    """Sitemap listing every community as a deep-linked dashboard URL."""
    s = db.get_settings()
    base = (s.get("site.canonicalUrl") or config.PUBLIC_BASE_URL or "").rstrip("/")
    records = _build_records()
    urls = [f"<url><loc>{base}/</loc><changefreq>weekly</changefreq><priority>1.0</priority></url>"]
    for r in records[:5000]:
        cid = r.get("id")
        if not cid:
            continue
        urls.append(f"<url><loc>{base}/#id={cid}</loc><changefreq>monthly</changefreq><priority>0.7</priority></url>")
    body = ('<?xml version="1.0" encoding="UTF-8"?>\n'
            '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
            + "\n".join(urls) + "\n</urlset>")
    return Response(content=body, media_type="application/xml")


@app.get("/cms")
@app.get("/cms/")
def cms_page() -> FileResponse:
    return FileResponse(CMS_ROOT / "index.html")


# Static mounts must come after explicit routes so the index resolves first.
app.mount("/cms", StaticFiles(directory=str(CMS_ROOT), html=True), name="cms")
app.mount("/", StaticFiles(directory=str(DASHBOARD_ROOT), html=True), name="dashboard")


def run() -> None:
    import uvicorn
    uvicorn.run("server.main:app", host=config.HOST, port=config.PORT, reload=False)


if __name__ == "__main__":
    run()
